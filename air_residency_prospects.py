"""
AiR / Residency donor-prospect list.

Re-run after the 2026-05-19 ticket refresh, with DTH attendees flowing in
through the freshly-loaded DataMerge. Target event set: all historical
AiR-tagged events + all 3 DTH events + ASE (the residency series).

Everett Score (matching the March 2026 methodology):
  40% target event attendance  (count of target events the patron attended)
  30% Contemporary + Dance genre score (avg of ContemporaryScore, DanceScore)
  20% Regularity                (attendance frequency, normalized)
  10% Avg Yearly Monetary       (financial capacity proxy, normalized)

Output: donors/AiR_Residency_Prospects.xlsx
"""
import os
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

os.chdir("/Users/antho/Documents/WPI-MW")

OUT = "donors/AiR_Residency_Prospects.xlsx"

NAVY        = "1A3A5C"
NAVY_LIGHT  = "E8EDF2"
WHITE       = "FFFFFF"
INK         = "1A2330"
INK_LT      = "5A6A7A"
ROW_STRIPE  = "F5F7FA"


def target_events(manifest):
    """Build the target event set: AiR-tagged + DTH + ASE."""
    m = manifest[manifest["EventType"] == "Live"].drop_duplicates("EventName").copy()
    air = m[m["EventLoB"] == "AiR"]
    dth = m[m["EventName"].str.contains("Dance Theatre|Dance Theater",
                                        case=False, na=False)]
    ase = m[m["EventName"].str.contains("Spiritual Ensemble|ASE",
                                        case=False, na=False)]
    targets = pd.concat([air, dth, ase]).drop_duplicates("EventName")
    return sorted(targets["EventName"].tolist())


def load_donor_data():
    """Combine all Donor_Classification tranches + compute MaxDonation."""
    path = "donors/Donor_Classification.xlsx"
    tranche_sheets = [
        "Major Patrons", "Growth Prospects", "Active Donors - Renew",
        "Dormant Donors - Reactivate", "Prime Non-Donor Prospects",
        "Lapsed Donors - Review",
    ]
    frames = []
    for sn in tranche_sheets:
        df = pd.read_excel(path, sheet_name=sn)
        df["Donor Tranche"] = sn
        frames.append(df)
    donors = pd.concat(frames, ignore_index=True)
    donors = donors.drop_duplicates("Account Name", keep="first")

    # MaxDonation from DonationsLatest
    dl = pd.read_excel("DonationsLatest.xlsx")
    amount_col = next((c for c in dl.columns if "amount" in c.lower()), None)
    account_col = next((c for c in dl.columns if c.lower() in (
        "account name", "accountname", "account")), None)
    if amount_col and account_col:
        max_g = dl.groupby(account_col)[amount_col].max().reset_index()
        max_g.columns = ["Account Name", "Max Gift"]
        donors = donors.merge(max_g, on="Account Name", how="left")
    else:
        donors["Max Gift"] = None

    return donors[["Account Name", "Donor Tranche", "Propensity Score",
                   "Max Gift", "Average Donation", "Lifetime Donations"]]


def target_attendance(merged, targets):
    """For each patron: # target events attended, comma-list of which ones,
    plus DTH-specific count to surface the freshest signal."""
    paid = merged[merged["EventName"].isin(targets)
                  & merged["Quantity"].gt(0)
                  & merged["TicketStatus"].eq("Active")].copy()
    by_acct = paid.groupby("AccountName")["EventName"].agg(
        lambda s: sorted(s.unique())
    ).reset_index()
    by_acct["# Target Events"] = by_acct["EventName"].str.len()
    by_acct["Target Events Attended"] = by_acct["EventName"].apply(", ".join)

    dth_events = [t for t in targets
                  if "Dance Theatre" in t or "Dance Theater" in t]
    by_acct["# DTH"] = by_acct["EventName"].apply(
        lambda ev: sum(1 for e in ev if e in dth_events))
    by_acct["DTH 2026"] = by_acct["EventName"].apply(
        lambda ev: "Yes" if "Dance Theatre of Harlem 2026" in ev else "")

    by_acct = by_acct.rename(columns={"AccountName": "Account Name"})
    return by_acct[["Account Name", "# Target Events", "Target Events Attended",
                    "# DTH", "DTH 2026"]]


def total_events_attended(merged):
    """Per-patron total event count (any active paid)."""
    paid = merged[merged["Quantity"].gt(0)
                  & merged["TicketStatus"].eq("Active")].copy()
    out = paid.groupby("AccountName").size().reset_index()
    out.columns = ["Account Name", "Events Attended"]
    return out


def minmax(s):
    """Min-max normalize a Series to [0, 100]."""
    s = pd.to_numeric(s, errors="coerce")
    lo, hi = s.min(), s.max()
    if pd.isna(hi) or hi == lo:
        return pd.Series([0.0] * len(s), index=s.index)
    return (s - lo) / (hi - lo) * 100


def _style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(bold=True, color=WHITE, size=11)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 30


def _stripe(ws, first_row, last_row, ncols):
    fill = PatternFill("solid", fgColor=ROW_STRIPE)
    for r in range(first_row, last_row + 1):
        if (r - first_row) % 2 == 1:
            for c in range(1, ncols + 1):
                cell = ws.cell(row=r, column=c)
                if cell.fill.start_color.rgb in (None, "00000000", "FFFFFFFF"):
                    cell.fill = fill


def _autowidth(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def main():
    print("Loading inputs…")
    patrons = pd.read_csv("Patrons.csv", low_memory=False)
    merged  = pd.read_csv("DataMerge.csv", low_memory=False)
    manifest = pd.read_excel("EventManifest.xlsx")
    donors = load_donor_data()
    print(f"  patrons: {len(patrons):,}  ·  donor rows: {len(donors):,}")

    targets = target_events(manifest)
    print(f"  {len(targets)} target events:")
    for t in targets: print(f"    · {t}")

    # Per-patron rollups
    ta = target_attendance(merged, targets)
    te = total_events_attended(merged)

    # Join everything to patrons (the master list)
    df = patrons.merge(donors, left_on="AccountName", right_on="Account Name", how="left")
    df = df.merge(ta, left_on="AccountName", right_on="Account Name", how="left", suffixes=("", "_ta"))
    df = df.merge(te, left_on="AccountName", right_on="Account Name", how="left", suffixes=("", "_te"))
    df["# Target Events"]        = df["# Target Events"].fillna(0).astype(int)
    df["Target Events Attended"] = df["Target Events Attended"].fillna("")
    df["# DTH"]                  = df["# DTH"].fillna(0).astype(int)
    df["DTH 2026"]               = df["DTH 2026"].fillna("")
    df["Events Attended"]        = df["Events Attended"].fillna(0).astype(int)

    # Score components — all normalized to [0, 100]
    df["GenreScore_CD"] = (pd.to_numeric(df["ContemporaryScore"], errors="coerce").fillna(0)
                           + pd.to_numeric(df["DanceScore"], errors="coerce").fillna(0)) / 2
    target_norm    = minmax(df["# Target Events"])
    genre_norm     = minmax(df["GenreScore_CD"])
    regularity_norm = minmax(df["Regularity"])
    aym_norm       = minmax(df["AYM"])

    df["Everett Score"] = (
        0.40 * target_norm
        + 0.30 * genre_norm
        + 0.20 * regularity_norm
        + 0.10 * aym_norm
    )

    # Affinity Signal — primary reason patron surfaces in this list
    def affinity(r):
        if r["# Target Events"] > 0:
            return "Attended Target Event"
        if r["GenreScore_CD"] >= 0.20:
            return "Genre Match (Contemporary/Dance)"
        if r["Regularity"] and r["Regularity"] >= 0.50:
            return "Regular Patron"
        return "Other"
    df["Affinity Signal"] = df.apply(affinity, axis=1)

    # Filter: focus list — patrons who EITHER attended a target event OR are
    # in any donor tranche (matches the spirit of the March 2026 list).
    has_target  = df["# Target Events"] > 0
    has_tranche = df["Donor Tranche"].fillna("").str.len() > 0
    df = df[has_target | has_tranche].copy()
    df = df.sort_values("Everett Score", ascending=False).reset_index(drop=True)

    # Build output frame — DTH columns added near target events for visibility
    out = pd.DataFrame({
        "Account Name":             df["AccountName"],
        "Everett Score":            df["Everett Score"].round(1),
        "Donor Tranche":            df["Donor Tranche"].fillna(""),
        "Patron Segment":           df["Segment"].fillna(""),
        "Affinity Signal":          df["Affinity Signal"],
        "Propensity Score":         df["Propensity Score"].round(1) if "Propensity Score" in df.columns else None,
        "Max Gift":                 df["Max Gift"],
        "Avg Yearly Spend":         df["AYM"],
        "Preferred Genre":          df["PreferredEventGenre"].fillna(""),
        "Contemporary":             pd.to_numeric(df["ContemporaryScore"], errors="coerce").round(3),
        "Dance":                    pd.to_numeric(df["DanceScore"], errors="coerce").round(3),
        "Choral":                   pd.to_numeric(df["ChoralScore"], errors="coerce").round(3),
        "# Target Events":          df["# Target Events"],
        "# DTH":                    df["# DTH"],
        "DTH 2026":                 df["DTH 2026"],
        "Target Events Attended":   df["Target Events Attended"],
        "Events Attended":          df["Events Attended"],
        "Regularity":               pd.to_numeric(df["Regularity"], errors="coerce").round(2),
        "Full Price Rate":          pd.to_numeric(df["FullPriceRate"], errors="coerce").round(2),
        "Region":                   df["RegionAssignment"].fillna(""),
        "Subscriber":               df["Subscriber"].fillna(""),
        "Patron Status":            df["PatronStatus"].fillna(""),
        "Email":                    df["Email"].fillna(""),
    })

    # Write formatted workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "AiR Residency Prospects"

    # Title + methodology subtitle
    ws.cell(row=1, column=1,
            value="Artist in Residence — Donor Prospects (refreshed 2026-05-19)"
            ).font = Font(bold=True, size=14, color=NAVY)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    ws.cell(row=2, column=1, value=(
        f"{len(out):,} patrons (target-event attenders + donors) sorted by Everett Score  ·  "
        f"Everett Score: 40% target event attendance + 30% Contemporary/Dance genre + "
        f"20% Regularity + 10% Avg Yearly Spend  ·  "
        f"Target events ({len(targets)}): AiR-tagged + DTH (all 3 seasons) + ASE  ·  "
        f"DTH-attender columns added to surface freshest signal from 2026-05-09"
    )).font = Font(italic=True, color=INK_LT, size=9.5)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=23)
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 32

    # Header
    for c, h in enumerate(out.columns, start=1):
        ws.cell(row=4, column=c, value=h)
    _style_header(ws, 4, len(out.columns))

    # Data rows
    int_fmt   = "#,##0"
    money_fmt = '"$"#,##0;[Red]-"$"#,##0;0'
    decimal2  = "0.00"
    decimal3  = "0.000"
    score_fmt = "0.0"
    for i, (_, r) in enumerate(out.iterrows()):
        row = 5 + i
        for c, col in enumerate(out.columns, start=1):
            v = r[col]
            if pd.isna(v): v = None
            cell = ws.cell(row=row, column=c, value=v)
            if col == "Everett Score":
                cell.number_format = score_fmt
            elif col == "Propensity Score":
                cell.number_format = score_fmt
            elif col in ("Max Gift", "Avg Yearly Spend"):
                cell.number_format = money_fmt
            elif col in ("Contemporary", "Dance", "Choral"):
                cell.number_format = decimal3
            elif col in ("Regularity", "Full Price Rate"):
                cell.number_format = decimal2
            elif col in ("# Target Events", "# DTH", "Events Attended"):
                cell.number_format = int_fmt

    last_row = 4 + len(out)
    _stripe(ws, 5, last_row, len(out.columns))
    ws.auto_filter.ref = f"A4:{get_column_letter(len(out.columns))}{last_row}"
    ws.freeze_panes = "C5"
    _autowidth(ws, [34, 11, 22, 16, 28, 12, 11, 11, 14, 12, 12, 12,
                    13, 7, 9, 80, 12, 11, 12, 14, 11, 12, 32])
    ws.sheet_view.showGridLines = False

    wb.save(OUT)
    print()
    print(f"  ✓ {OUT}")
    print(f"  {len(out):,} prospects  ·  top tier counts:")
    print(out["Donor Tranche"].value_counts(dropna=False).head(8).to_string())
    print(f"  Affinity signal counts:")
    print(out["Affinity Signal"].value_counts().to_string())


if __name__ == "__main__":
    main()
