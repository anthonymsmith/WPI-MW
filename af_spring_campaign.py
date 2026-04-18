"""
AF_Spring_Campaign.py — Annual Fund Spring Campaign Prospect List

Sources:
  FY donation review Annual Fund *.xlsx  — AF transactions with Donation Record Type
  Patrons.csv                            — patron engagement metrics

Segments (ordered by spring campaign priority):
  S1  Renew — Spring/Both donors    gave in FY25 or FY26, spring-leaning giving season
  S2  Renew — Fall donors           gave in FY25 or FY26, fall-leaning giving season
  S3  Reactivate                    gave to AF before FY25, now lapsed
  S4  Acquire — Ticket Donors       PatronTicket only; engaged but never cash-gifted
  S5  Acquire — Prospects           no AF history; high-engagement patrons

Output: AF_Spring_Campaign.xlsx
"""
import glob
import math
import os
import re
from datetime import datetime

import numpy as np
import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ── Config ────────────────────────────────────────────────────────────────────
DATA_DIR   = '/Users/antho/Documents/WPI-MW'
OUTPUT     = os.path.join(DATA_DIR, 'AF_Spring_Campaign.xlsx')
TODAY      = datetime.today()
CURRENT_FY = 2025   # FY2025-2026 (Jul 2025 – Jun 2026)

# FY helpers
def close_date_fy(dt):
    return dt.year if dt.month >= 7 else dt.year - 1

def fy_label(fy):
    return f"FY{fy}-{fy+1 - 2000:02d}"


# ── 1. Load Annual Fund file ──────────────────────────────────────────────────
af_files = sorted(glob.glob(os.path.join(DATA_DIR, 'FY donation review Annual Fund*.xlsx')))
assert af_files, "No Annual Fund file found in DATA_DIR"
af_file = af_files[-1]   # use most recent
print(f"Loading: {os.path.basename(af_file)}")

wb = openpyxl.load_workbook(af_file, data_only=True)
ws = wb.active

# Header is row 14; data starts row 15
raw = []
current_name = ''
for r in ws.iter_rows(min_row=15, values_only=True):
    name, acct_id, amt, rec_type, close_date = r[1], r[3], r[8], r[9], r[10]
    if not rec_type:          # skip Subtotal / blank rows
        continue
    # Forward-fill AccountName: SF report only puts name on the first gift row per account
    row_name = str(name).strip() if name else ''
    if row_name and row_name != 'Subtotal':
        current_name = row_name
    raw.append({
        'AccountName': current_name,
        'AccountID':   str(acct_id).strip() if acct_id else '',
        'Amount':      float(amt) if amt else 0.0,
        'RecordType':  rec_type,
        'CloseDate':   close_date,
    })

af = pd.DataFrame(raw)
af = af[af['AccountName'] != '']

# Parse close dates
af['CloseDate'] = pd.to_datetime(af['CloseDate'], errors='coerce', format='mixed')
af = af.dropna(subset=['CloseDate'])
af['FY'] = af['CloseDate'].apply(close_date_fy)
af['Month'] = af['CloseDate'].dt.month
af['Season'] = af['Month'].apply(lambda m: 'Spring' if 1 <= m <= 6 else 'Fall')

print(f"  {len(af):,} donation rows | "
      f"{af['AccountName'].nunique():,} unique accounts | "
      f"FY range {af['FY'].min()}-{af['FY'].max()}")


# ── 2. Compute per-account AF metrics ─────────────────────────────────────────
cash = af[af['RecordType'] == 'Donation'].copy()
ticket = af[af['RecordType'] == 'PatronTicket Donation'].copy()

# --- Cash donor metrics ---
def giving_streak(fy_series):
    """Consecutive FYs with a gift ending at the most recent FY with a gift."""
    fys = sorted(fy_series.unique())
    if not fys:
        return 0
    streak, latest = 1, fys[-1]
    for fy in reversed(fys[:-1]):
        if latest - fy == 1:
            streak += 1
            latest = fy
        else:
            break
    return streak

def giving_season(season_series):
    fall   = (season_series == 'Fall').sum()
    spring = (season_series == 'Spring').sum()
    if fall == 0 and spring == 0:
        return 'Unknown'
    if fall >= spring * 1.5:
        return 'Fall'
    if spring >= fall * 1.5:
        return 'Spring'
    return 'Both'

cash_agg = cash.groupby('AccountName').agg(
    AFCashGiftCount   = ('Amount', 'count'),
    AFLifetimeCash    = ('Amount', 'sum'),
    AFMaxCashGift     = ('Amount', 'max'),
    AFAvgCashGift     = ('Amount', 'mean'),
    AFLastCashDate    = ('CloseDate', 'max'),
    AFFirstCashDate   = ('CloseDate', 'min'),
).reset_index()

# Compute FY-dependent metrics separately and join on AccountName to avoid index misalignment
streak_s  = cash.groupby('AccountName')['FY'].apply(giving_streak).rename('AFGivingStreak')
season_s  = cash.groupby('AccountName')['Season'].apply(giving_season).rename('AFGivingSeason')
gave_fy25    = cash.groupby('AccountName')['FY'].apply(lambda x: (x == 2024).any()).rename('AFGaveFY25')
gave_fy26    = cash.groupby('AccountName')['FY'].apply(lambda x: (x == CURRENT_FY).any()).rename('AFGaveFY26')
last_gift_fy = cash.groupby('AccountName')['FY'].apply(
    lambda x: fy_label(x.max())
).rename('AFLastGiftFY')
def streak_fys(fy_series):
    """Consecutive FYs ending at the most recent gift — matches AFGivingStreak."""
    fys = sorted(fy_series.unique())
    if not fys:
        return ''
    run, latest = [fys[-1]], fys[-1]
    for fy in reversed(fys[:-1]):
        if latest - fy == 1:
            run.append(fy)
            latest = fy
        else:
            break
    return ', '.join(fy_label(fy) for fy in sorted(run))

gift_fys = cash.groupby('AccountName')['FY'].apply(streak_fys).rename('AFGiftFYs')

for s in (streak_s, season_s, gave_fy25, gave_fy26, last_gift_fy, gift_fys):
    cash_agg = cash_agg.merge(s.reset_index(), on='AccountName', how='left')

cash_agg['AFRecentDonor'] = cash_agg['AFGaveFY25'] | cash_agg['AFGaveFY26']
cash_agg['AFMonthsSinceLastCash'] = (
    (TODAY - cash_agg['AFLastCashDate']).dt.days / 30.44
).round(1)
cash_agg['AFIsCashDonor']  = True

# --- Ticket-only donors ---
ticket_names = set(ticket['AccountName'].unique()) - set(cash_agg['AccountName'].unique())
ticket_only  = pd.DataFrame({'AccountName': sorted(ticket_names), 'AFIsTicketDonor': True})

print(f"  Cash donors: {len(cash_agg):,} | Ticket-only: {len(ticket_only):,}")


# ── 3. Load Patrons.csv ───────────────────────────────────────────────────────
patrons = pd.read_csv(os.path.join(DATA_DIR, 'Patrons.csv'), low_memory=False)
print(f"Patrons.csv: {len(patrons):,} patrons")

# ── 3b. Load DonationsLatest.xlsx — all-funds giving ─────────────────────────
all_don = pd.read_excel(os.path.join(DATA_DIR, 'DonationsLatest.xlsx'))
all_don['Close Date'] = pd.to_datetime(all_don['Close Date'], errors='coerce')
all_funds = all_don.groupby('Account Name').agg(
    AllFundsLifetime  = ('Amount', 'sum'),
    AllFundsMaxGift   = ('Amount', 'max'),
    AllFundsGiftCount = ('Amount', 'count'),
    AllFundsLastDate  = ('Close Date', 'max'),
).reset_index().rename(columns={'Account Name': 'AccountName'})
all_funds['AllFundsDonor'] = True

# Detect named campaign participation per account
def detect_campaigns(donation_names):
    text = ' '.join(str(n).lower() for n in donation_names if pd.notna(n))
    tags = []
    if 'chorus' in text:                                    tags.append('Chorus')
    if 'campaign 2025' in text or 'campaign 25' in text:   tags.append('Campaign 25')
    if 'sponsor' in text or 'concert fund' in text:        tags.append('Concert Sponsor')
    if 'endow' in text:                                     tags.append('Endowment')
    return ', '.join(tags)

other_camps = (
    all_don.groupby('Account Name')['Donation Name']
    .apply(detect_campaigns)
    .rename('OtherCampaigns')
    .reset_index()
    .rename(columns={'Account Name': 'AccountName'})
)
all_funds = all_funds.merge(other_camps, on='AccountName', how='left')
all_funds['OtherCampaigns'] = all_funds['OtherCampaigns'].fillna('')
print(f"DonationsLatest.xlsx: {len(all_don):,} records | {len(all_funds):,} unique accounts")


# ── 4. Join ───────────────────────────────────────────────────────────────────
df = patrons.copy()

# Deduplicate Patrons.csv on AccountName (keep highest Frequency row per name)
dupes = df['AccountName'].duplicated(keep=False).sum()
if dupes:
    print(f"  Deduplicating: {dupes} rows share an AccountName — keeping highest-Frequency record")
    df = df.sort_values('Frequency', ascending=False).drop_duplicates('AccountName', keep='first')

df = df.merge(cash_agg,   on='AccountName', how='left')
df = df.merge(ticket_only, on='AccountName', how='left')
df = df.merge(all_funds,   on='AccountName', how='left')

bool_cols = ['AFIsCashDonor', 'AFIsTicketDonor', 'AFRecentDonor',
             'AFGaveFY25', 'AFGaveFY26', 'AllFundsDonor']
for c in bool_cols:
    df[c] = df[c].infer_objects(copy=False).fillna(False)

df[['AFCashGiftCount','AFLifetimeCash','AFMaxCashGift','AFAvgCashGift',
    'AFGivingStreak','AllFundsLifetime','AllFundsMaxGift','AllFundsGiftCount']] = (
    df[['AFCashGiftCount','AFLifetimeCash','AFMaxCashGift','AFAvgCashGift',
        'AFGivingStreak','AllFundsLifetime','AllFundsMaxGift','AllFundsGiftCount']].fillna(0)
)
df['AFGivingSeason'] = df['AFGivingSeason'].fillna('None')
df['AFLastGiftFY']   = df['AFLastGiftFY'].fillna('')
df['AFGiftFYs']      = df['AFGiftFYs'].fillna('')

# Flag patrons who give to other campaigns but not AF cash
df['OtherCampaignDonor'] = df['AllFundsDonor'] & ~df['AFIsCashDonor']
df['OtherCampaigns']     = df['OtherCampaigns'].fillna('')

matched_cash   = df['AFIsCashDonor'].sum()
matched_ticket = df['AFIsTicketDonor'].sum()
other_campaign = df['OtherCampaignDonor'].sum()
print(f"Joined: {matched_cash} AF cash donors | {matched_ticket} ticket-only | "
      f"{other_campaign} give to other campaigns but not AF")


# ── 5. Segment assignment ─────────────────────────────────────────────────────
def assign_segment(row):
    if row['AFIsCashDonor']:
        if row['AFRecentDonor']:
            if row['AFGivingSeason'] in ('Spring', 'Both'):
                return 'S1'   # Renew — Spring/Both
            else:
                return 'S2'   # Renew — Fall
        else:
            return 'S3'       # Reactivate
    elif row['AFIsTicketDonor']:
        return 'S4'           # Acquire — Ticket Donors
    else:
        return 'S5'           # Acquire — Prospects

df['Segment_Code'] = df.apply(assign_segment, axis=1)

seg_labels = {
    'S1': 'Renew — Spring/Both',
    'S2': 'Renew — Fall',
    'S3': 'Reactivate',
    'S4': 'Acquire — Ticket Donors',
    'S5': 'Acquire — Prospects',
}
df['CampaignSegment'] = df['Segment_Code'].map(seg_labels).fillna('Excluded — below threshold')

for code, label in seg_labels.items():
    print(f"  {code} {label}: {(df['Segment_Code'] == code).sum()}")


# ── 6. Filter S5 and score all segments ───────────────────────────────────────
def minmax(series):
    mn, mx = series.min(), series.max()
    if mx == mn:
        return pd.Series(0.5, index=series.index)
    return (series - mn) / (mx - mn)

# S5: filter to patrons with meaningful relationship depth before scoring
# Regularity >= 0.3 (consistent attenders) AND Lifespan >= 3 years (established relationship)
s5_mask = df['Segment_Code'] == 'S5'
s5_filter = s5_mask & (df['Regularity'] >= 0.3) & (df['Lifespan'] >= 3.0)
df.loc[s5_mask & ~s5_filter, 'Segment_Code'] = 'S5_excluded'
print(f"  S5 after filter (Regularity>=0.3, Lifespan>=3y): "
      f"{s5_filter.sum()} kept / {s5_mask.sum() - s5_filter.sum()} excluded")

# Donor score (S1-S3): gift history + engagement
donor_mask = df['Segment_Code'].isin(['S1', 'S2', 'S3'])
df.loc[donor_mask, 'AFScore'] = (
    0.35 * minmax(df.loc[donor_mask, 'AFMaxCashGift']) +
    0.25 * minmax(df.loc[donor_mask, 'AFGivingStreak']) +
    0.25 * minmax(df.loc[donor_mask, 'AYM']) +
    0.15 * minmax(df.loc[donor_mask, 'Regularity'])
)

# S4 score: ticket-return donors — capacity + relationship depth
s4_mask = df['Segment_Code'] == 'S4'
df.loc[s4_mask, 'AFScore'] = (
    0.40 * minmax(df.loc[s4_mask, 'AYM']) +
    0.30 * minmax(df.loc[s4_mask, 'Regularity']) +
    0.20 * minmax(df.loc[s4_mask, 'RFMScore']) +
    0.10 * minmax(df.loc[s4_mask, 'FullPriceRate'])
)

# S5 score: no donation history — relationship depth + growth trajectory
s5_active = df['Segment_Code'] == 'S5'
df.loc[s5_active, 'AFScore'] = (
    0.35 * minmax(df.loc[s5_active, 'Regularity']) +
    0.25 * minmax(df.loc[s5_active, 'Lifespan']) +
    0.25 * minmax(df.loc[s5_active, 'GrowthScore']) +
    0.15 * minmax(df.loc[s5_active, 'AYM'])
)

df['AFScore'] = (df['AFScore'] * 100).round(1)

# ── 6b. Donor tier — uses max gift across AF + all funds ─────────────────────
# Beethoven Society = MW's ~$1,200 giving level; Major = meaningfully above that
BEETHOVEN_MIN = 1000
MAJOR_MIN     = 2500

def donor_tier(row):
    # Use the higher of AF cash gift or all-funds gift for tier placement
    max_gift = max(float(row.get('AFMaxCashGift', 0) or 0),
                   float(row.get('AllFundsMaxGift', 0) or 0))
    if max_gift == 0:
        if row.get('OtherCampaignDonor', False):
            return 'Other Campaign'
        return ''
    if max_gift >= MAJOR_MIN:
        return 'Major ($2.5K+)'
    if max_gift >= BEETHOVEN_MIN:
        return 'Beethoven ($1K–$2.4K)'
    if max_gift >= 250:
        return 'Mid ($250–$999)'
    return 'Small (<$250)'

df['DonorTier'] = df.apply(donor_tier, axis=1)

tier_order = {
    'Major ($2.5K+)': 0, 'Beethoven ($1K–$2.4K)': 1,
    'Mid ($250–$999)': 2, 'Small (<$250)': 3,
    'Other Campaign': 4, '': 5,
}
for tier in tier_order:
    n = (df['DonorTier'] == tier).sum()
    if n: print(f"  Donor tier {tier}: {n}")


# ── 7. Build output columns ───────────────────────────────────────────────────
df['AFLastCashDate_str'] = df['AFLastCashDate'].dt.strftime('%Y-%m-%d').fillna('')

out_cols = [
    'AccountName',
    'CampaignSegment', 'AFScore', 'DonorTier',
    'Subscriber', 'ChorusMember', 'PatronStatus',
    'AFGivingSeason', 'AFMaxCashGift', 'AFLastGiftFY', 'AFGivingStreak',
    'AllFundsMaxGift', 'AllFundsGiftCount', 'OtherCampaigns',
    'AYM', 'Frequency', 'Regularity', 'RFMScore', 'FullPriceRate',
    'Segment', 'LatestSeason', 'RegionAssignment',
    'Email',
]

display_names = {
    'AccountName':       'Account Name',
    'CampaignSegment':   'Campaign Segment',
    'AFScore':           'AF Score',
    'DonorTier':         'Donor Tier',
    'Subscriber':        'Subscriber',
    'ChorusMember':      'In Chorus',
    'PatronStatus':      'Patron Status',
    'AFGivingSeason':    'AF Giving Season',
    'AFMaxCashGift':     'AF Max Gift',
    'AFLastGiftFY':      'Last Gift FY',
    'AFGiftFYs':         'Gift Streak FYs',
    'AFGivingStreak':    'Streak (FYs)',
    'AllFundsMaxGift':   'All-Funds Max Gift',
    'AllFundsGiftCount': 'All-Funds Gifts',
    'OtherCampaigns':    'Other Campaigns',
    'AYM':               'Avg Yearly Spend',
    'Frequency':         'Events Attended',
    'Regularity':        'Regularity',
    'RFMScore':          'RFM Score',
    'FullPriceRate':     'Full Price Rate',
    'Segment':           'Patron Segment',
    'LatestSeason':      'Latest Season',
    'RegionAssignment':  'Region',
    'Email':             'Email',
}


# ── 8. Write Excel ────────────────────────────────────────────────────────────
NAVY   = '1A3A5C'
WHITE  = 'FFFFFF'
TEAL   = '2A9EA0'
GREEN  = 'D5E8D4'   # S1 Renew Spring
BLUE   = 'DAE8FC'   # S2 Renew Fall
AMBER  = 'FFE6CC'   # S3 Reactivate
PURPLE = 'E1D5E7'   # S4 Ticket donors
GRAY   = 'F5F5F5'   # S5 Prospects

seg_colors = {
    'S1': GREEN, 'S2': BLUE, 'S3': AMBER, 'S4': PURPLE, 'S5': GRAY,
}

def cfill(h): return PatternFill('solid', fgColor=h)

wb_out = openpyxl.Workbook()

# ── Overview sheet ──
ws_ov = wb_out.active
ws_ov.title = 'Overview'
ws_ov.sheet_view.zoomScale = 125

ws_ov.merge_cells('A1:D1')
c = ws_ov['A1']
c.value = 'Annual Fund — Spring Campaign Prospect List'
c.font  = Font(bold=True, color=WHITE, size=14)
c.fill  = cfill(NAVY)
c.alignment = Alignment(horizontal='center', vertical='center')
ws_ov.row_dimensions[1].height = 26

ws_ov['A2'].value = f'Generated {TODAY.strftime("%Y-%m-%d")} · Patrons.csv + {os.path.basename(af_file)}'
ws_ov['A2'].font  = Font(italic=True, size=9, color='666666')

# Outreach split: Personal Calls = S1 all + S2/S3 Major & Beethoven
#                 Email & Letter = S2/S3 Mid & Small + S4 all
MAJOR_BEETHOVEN_TIERS = {'Major ($2.5K+)', 'Beethoven ($1K–$2.4K)'}

s1_mask  = df['Segment_Code'] == 'S1'
s2_mask  = df['Segment_Code'] == 'S2'
s3_mask  = df['Segment_Code'] == 'S3'
s4_mask2 = df['Segment_Code'] == 'S4'

personal_calls_mask = (
    s1_mask |
    (s2_mask & df['DonorTier'].isin(MAJOR_BEETHOVEN_TIERS)) |
    (s3_mask & df['DonorTier'].isin(MAJOR_BEETHOVEN_TIERS))
)
email_letter_mask = (
    (s2_mask & ~df['DonorTier'].isin(MAJOR_BEETHOVEN_TIERS)) |
    (s3_mask & ~df['DonorTier'].isin(MAJOR_BEETHOVEN_TIERS)) |
    s4_mask2
)

n_s1       = int(s1_mask.sum())
n_s2_mb    = int((s2_mask & df['DonorTier'].isin(MAJOR_BEETHOVEN_TIERS)).sum())
n_s3_mb    = int((s3_mask & df['DonorTier'].isin(MAJOR_BEETHOVEN_TIERS)).sum())
n_personal = int(personal_calls_mask.sum())
n_s2_ms    = int((s2_mask & ~df['DonorTier'].isin(MAJOR_BEETHOVEN_TIERS)).sum())
n_s3_ms    = int((s3_mask & ~df['DonorTier'].isin(MAJOR_BEETHOVEN_TIERS)).sum())
n_s4       = int(s4_mask2.sum())
n_email    = int(email_letter_mask.sum())

r = 4
for col, h in enumerate(['Segment', 'Label', 'Count', 'Notes'], 1):
    c = ws_ov.cell(row=r, column=col, value=h)
    c.font = Font(bold=True, color=WHITE, size=10)
    c.fill = cfill(TEAL)
r += 1

def ov_section(label, fill_hex, text_color=WHITE):
    global r
    ws_ov.merge_cells(f'A{r}:D{r}')
    c = ws_ov.cell(row=r, column=1, value=label)
    c.font = Font(bold=True, color=text_color, size=10)
    c.fill = cfill(fill_hex)
    r += 1

def ov_row(seg_code, label, count, note, color):
    global r
    for col_i, val in enumerate([seg_code, label, int(count), note], 1):
        c = ws_ov.cell(row=r, column=col_i, value=val)
        c.fill = cfill(color)
        if col_i == 4:
            c.font = Font(italic=True, size=9)
    r += 1

def ov_total(label, count):
    global r
    ws_ov.cell(row=r, column=1, value=label).font = Font(bold=True, size=10)
    ws_ov.cell(row=r, column=3, value=int(count)).font = Font(bold=True, size=10)
    r += 2

# ── Personal Calls section ──
ov_section('PERSONAL OUTREACH — Priority for board & committee outreach', NAVY)
ov_row('S1', 'Renew — Spring/Both', n_s1,
       'Recent AF donors, spring-aligned giving season — highest priority', GREEN)
ov_row('S2 — Major & Beethoven', 'Renew — Fall (Major & Beethoven)', n_s2_mb,
       'Recent AF donors, fall-leaning — ask now, may renew off-cycle', BLUE)
ov_row('S3 — Major & Beethoven', 'Reactivate (Major & Beethoven)', n_s3_mb,
       'AF donors not seen in FY25/26 — reactivation ask', AMBER)
ov_total('Total personal outreach', n_personal)

# ── Email & Letter section ──
ov_section('EMAIL & LETTER — Extended outreach list', TEAL)
ov_row('S2 — Mid & Small', 'Renew — Fall (Mid & Small)', n_s2_ms,
       'Recent AF donors, fall-leaning — ask now, may renew off-cycle', BLUE)
ov_row('S3 — Mid & Small', 'Reactivate (Mid & Small)', n_s3_ms,
       'AF donors not seen in FY25/26 — reactivation ask', AMBER)
ov_row('S4', 'Acquire — Ticket Donors', n_s4,
       'Engaged but never a cash gift — first-gift ask', PURPLE)
ov_total('Total email & letter', n_email)

# ── Reference section ──
ov_section('FOR REFERENCE ONLY — Not an outreach list', 'FFF2CC', text_color='885500')
s5_cnt = int((df['Segment_Code'] == 'S5').sum())
ov_row('S5', seg_labels['S5'], s5_cnt,
       'No AF history; Regularity≥30% + Lifespan≥3yr filter applied; ranked by relationship depth', GRAY)
r += 1

ws_ov.cell(row=r, column=1, value='Scoring approach:').font = Font(bold=True, size=9)
r += 1
ws_ov.cell(row=r, column=1,
    value='S1–S3 (donors):    35% Max Cash Gift + 25% Giving Streak + 25% Avg Yearly Spend + 15% Regularity')
ws_ov.cell(row=r, column=1).font = Font(italic=True, size=9)
r += 1
ws_ov.cell(row=r, column=1,
    value='S4 (ticket donors): 40% Avg Yearly Spend + 30% Regularity + 20% RFM Score + 10% Full Price Rate')
ws_ov.cell(row=r, column=1).font = Font(italic=True, size=9)
r += 1
ws_ov.cell(row=r, column=1,
    value='S5 (reference):     35% Regularity + 25% Lifespan + 25% Growth Score + 15% Avg Yearly Spend')
ws_ov.cell(row=r, column=1).font = Font(italic=True, size=9)

for col, w in zip('ABCD', [22, 30, 8, 62]):
    ws_ov.column_dimensions[col].width = w


# ── Column definitions shared by Personal Calls and Email & Letter ─────────────
outreach_cols = [
    'AccountName',
    'CampaignSegment', 'DonorTier',
    'Subscriber', 'ChorusMember', 'PatronStatus',
    'AFGivingSeason', 'AFMaxCashGift', 'AFLastGiftFY', 'AFGivingStreak',
    'AllFundsMaxGift', 'OtherCampaigns',
    'AYM', 'Frequency', 'RegionAssignment',
    'Email',
]
outreach_display = {k: display_names[k] for k in outreach_cols}

# Column widths from user's manual adjustments to Priority Outreach
outreach_widths = {
    'AccountName': 28,
    'CampaignSegment': 14.67, 'DonorTier': 12.33,
    'Subscriber': 7.17, 'ChorusMember': 5.83, 'PatronStatus': 10,
    'AFGivingSeason': 7, 'AFMaxCashGift': 6.67, 'AFLastGiftFY': 8.5, 'AFGivingStreak': 4.67,
    'AllFundsMaxGift': 8.5, 'OtherCampaigns': 31.33,
    'AYM': 8.5, 'Frequency': 7.5, 'RegionAssignment': 10,
    'Email': 20.5,
}

def write_outreach_sheet(ws, src_df, title_text, zoom=125):
    """Write a formatted outreach sheet from src_df (rows already sorted)."""
    ws.sheet_view.zoomScale = zoom
    ws.merge_cells(f'A1:{get_column_letter(len(outreach_cols))}1')
    c = ws['A1']
    c.value = title_text
    c.font = Font(bold=True, color=WHITE, size=13)
    c.fill = cfill(NAVY)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 24

    for col_i, col_name in enumerate(outreach_cols, 1):
        c = ws.cell(row=2, column=col_i, value=outreach_display[col_name])
        c.font = Font(bold=True, color=WHITE, size=9)
        c.fill = cfill(TEAL)
        c.alignment = Alignment(horizontal='center', wrap_text=True)
    ws.row_dimensions[2].height = 28

    for row_i, (_, row) in enumerate(src_df[outreach_cols].iterrows(), 3):
        seg_code = 'S1' if 'Spring' in str(row['CampaignSegment']) else \
                   'S2' if 'Fall'   in str(row['CampaignSegment']) else \
                   'S3' if 'React'  in str(row['CampaignSegment']) else 'S4'
        row_color = seg_colors.get(seg_code, GRAY)
        for col_i, col_name in enumerate(outreach_cols, 1):
            val = row[col_name]
            if pd.isna(val):            val = ''
            elif isinstance(val, bool): val = 'Yes' if val else ''
            elif col_name in ('AYM', 'AFMaxCashGift', 'AllFundsMaxGift'):
                val = round(float(val), 0) if val != '' else 0
            c = ws.cell(row=row_i, column=col_i, value=val)
            c.fill = cfill(row_color)
            c.font = Font(size=9)
            if col_name in ('AYM', 'AFMaxCashGift', 'AllFundsMaxGift'):
                c.number_format = '"$"#,##0'
                c.alignment = Alignment(horizontal='right')

    for col_i, col_name in enumerate(outreach_cols, 1):
        ws.column_dimensions[get_column_letter(col_i)].width = outreach_widths.get(col_name, 10)
    ws.freeze_panes = 'A3'


# ── Personal Calls sheet (S1 all + S2/S3 Major & Beethoven) ──────────────────
pc_src = df[personal_calls_mask].copy()
pc_src['_tier_rank'] = pc_src['DonorTier'].map(tier_order).fillna(5)
pc_src = pc_src.sort_values(
    ['Segment_Code', '_tier_rank', 'AFScore'], ascending=[True, True, False]
).reset_index(drop=True)

ws_pc = wb_out.create_sheet(title='Personal Outreach', index=1)
write_outreach_sheet(ws_pc, pc_src,
    f'Annual Fund — Spring Campaign  |  Personal Outreach  ({n_personal} patrons)')


# ── Email & Letter sheet (S2/S3 Mid & Small + S4) ────────────────────────────
el_src = df[email_letter_mask].copy()
el_src['_tier_rank'] = el_src['DonorTier'].map(tier_order).fillna(5)
el_src = el_src.sort_values(
    ['Segment_Code', '_tier_rank', 'AFScore'], ascending=[True, True, False]
).reset_index(drop=True)

ws_el = wb_out.create_sheet(title='Email & Letter', index=2)
write_outreach_sheet(ws_el, el_src,
    f'Annual Fund — Spring Campaign  |  Email & Letter  ({n_email} patrons)')


# Column widths for detail sheets (matching outreach_widths where cols overlap)
detail_col_widths = {
    'AccountName': 28,
    'CampaignSegment': 14.67, 'AFScore': 7, 'DonorTier': 12.33,
    'Subscriber': 7.17, 'ChorusMember': 5.83, 'PatronStatus': 10,
    'AFGivingSeason': 7, 'AFMaxCashGift': 6.67, 'AFLastGiftFY': 8.5, 'AFGivingStreak': 4.67,
    'AFGiftFYs': 20,
    'AllFundsMaxGift': 8.5, 'AllFundsGiftCount': 7, 'OtherCampaigns': 20,
    'AYM': 8.5, 'Frequency': 7.5, 'Regularity': 8, 'RFMScore': 7, 'FullPriceRate': 8,
    'Segment': 12, 'LatestSeason': 9, 'RegionAssignment': 10,
    'Email': 20.5,
}

# ── One sheet per segment ──
for code, label in seg_labels.items():
    seg_df = df[df['Segment_Code'] == code][out_cols].copy()
    seg_df = seg_df.sort_values('AFScore', ascending=False).reset_index(drop=True)

    safe_title = label.replace('/', '-')[:31]
    ws = wb_out.create_sheet(title=safe_title)
    ws.sheet_view.zoomScale = 125
    color = seg_colors[code]

    # Title — flag S5 as reference only
    title_prefix = '[REFERENCE ONLY] ' if code == 'S5' else ''
    ws.merge_cells(f'A1:{get_column_letter(len(out_cols))}1')
    c = ws['A1']
    c.value = f'{title_prefix}{code} — {label}  ({len(seg_df)} prospects)'
    c.font  = Font(bold=True, color=WHITE, size=12)
    c.fill  = cfill(NAVY)
    c.alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 22

    # Column headers
    for col_i, col_name in enumerate(out_cols, 1):
        c = ws.cell(row=2, column=col_i, value=display_names[col_name])
        c.font = Font(bold=True, color=WHITE, size=9)
        c.fill = cfill(TEAL)
        c.alignment = Alignment(horizontal='center', wrap_text=True)
    ws.row_dimensions[2].height = 28

    # Data rows
    for row_i, (_, row) in enumerate(seg_df.iterrows(), 3):
        row_color = color if row_i % 2 == 1 else 'FFFFFF'
        for col_i, col_name in enumerate(out_cols, 1):
            val = row[col_name]
            # Clean up booleans and NaN
            if pd.isna(val):
                val = ''
            elif isinstance(val, bool):
                val = 'Yes' if val else ''
            elif col_name in ('AYM', 'AFMaxCashGift', 'AFLifetimeCash', 'AFAvgCashGift', 'AllFundsMaxGift'):
                val = round(float(val), 0) if val != '' else 0
            elif col_name == 'AFScore':
                val = round(float(val), 1) if val != '' else 0
            elif col_name in ('Regularity', 'FullPriceRate'):
                val = round(float(val), 3) if val != '' else 0

            c = ws.cell(row=row_i, column=col_i, value=val)
            c.fill = cfill(row_color)
            c.font = Font(size=9)

            # Number formats
            if col_name in ('AYM', 'AFMaxCashGift', 'AFLifetimeCash', 'AFAvgCashGift', 'AllFundsMaxGift'):
                c.number_format = '"$"#,##0'
                c.alignment = Alignment(horizontal='right')
            elif col_name in ('Regularity', 'FullPriceRate'):
                c.number_format = '0%'
                c.alignment = Alignment(horizontal='right')
            elif col_name == 'AFScore':
                c.number_format = '0.0'
                c.alignment = Alignment(horizontal='right')

    for col_i, col_name in enumerate(out_cols, 1):
        ws.column_dimensions[get_column_letter(col_i)].width = detail_col_widths.get(col_name, 10)

    ws.freeze_panes = 'A3'

wb_out.save(OUTPUT)
print(f"\n✓ Saved: {OUTPUT}")
print()
print("=== SEGMENT SUMMARY ===")
for code, label in seg_labels.items():
    sub = df[df['Segment_Code'] == code]
    cash_sub = sub[sub['AFIsCashDonor']]
    print(f"  {code}  {label:<30}  {len(sub):>5} patrons", end='')
    if len(cash_sub) > 0:
        print(f"   max gift ${cash_sub['AFMaxCashGift'].max():>8,.0f}"
              f"   avg gift ${cash_sub['AFAvgCashGift'].mean():>7,.0f}")
    else:
        print(f"   avg AYM ${sub['AYM'].mean():>8,.0f}")
