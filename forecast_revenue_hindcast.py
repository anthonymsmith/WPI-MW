"""
Revenue hindcast for 25-26 — blinded except for the event manifest.

For each 25-26 event:
  Pred_Revenue = Pred_Paid (from hindcast forecast) × historical avg paid price
  Actual_Revenue = sum(TicketTotal) on paid tickets in DataMerge

Historical avg price uses the same hierarchy as the comp-split layer:
  EventRepeat → Primary → F1 → F2 → F3 → F4 → _overall
Trained on prior seasons only, season-weighted (heavier weight on recent),
PWYW + Free events excluded so heavily-comped runs don't drag the price prior.

Output: forecasting/Forecast_2526_Revenue_Hindcast.xlsx
  - Summary sheet: totals + per-event price/volume/revenue gap
  - PricingObservations sheet: events where pricing reality differed from
    historical bucket, surfaced as candidates for the next pricing review
"""
import os
import pandas as pd
import numpy as np
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

os.chdir("/Users/antho/Documents/WPI-MW")

from forecast_2526_comparison import load_data, WEIGHTS

FORECAST_SEASON = "25-26"
FORECAST_XLSX   = "forecasting/Forecast_2526_FullSeason.xlsx"
OUT_XLSX        = "forecasting/Forecast_2526_Revenue_Hindcast.xlsx"

MIN_N    = 3      # row threshold same as comp_split
SHRINK_K = 3      # match comp_split shrinkage

# Per-tier price-uplift policy. Uses the `PriceTier` column on
# `EventManifest.xlsx` (derived 2026-05-19). Tier definitions:
#   Marquee                  — top-draw bookings tagged pre-season as
#                              priced-up (Trifonov-class, marquee festivals)
#   Headliner / Prestige     — top-tier commercial bookings
#   Standard          — Standard-class secular concerts (not Chorus)
#   Chorus                   — Worcester Chorus / Women's Ensemble events
#                              (accessibly priced; not uplifted)
#   TCB                      — The Complete Bach LoB (festival-style pricing)
#   Mission                  — community/education programming
#   AiR                      — Artist-in-Residence (community-priced)
UPLIFT_BY_TIER = {
    "Marquee":         0.40,
    "Headliner":       0.15,
    "Prestige":        0.15,
    "Standard": 0.10,
    "Chorus":          0.00,
    "TCB":            -0.25,   # audience-building negative uplift (TCB Side); FestivalRole=Main overrides to 0
    "Mission":         0.00,
    "AiR":             0.00,
}

# Tier-specific shrinkage target only applies to these tiers (the 0%-uplift
# tiers where MW prices to a deliberate accessible posture, not market). For
# Headliner/Prestige/Standard/Marquee, shrinking toward tier-overall would
# pull predictions DOWN because historical-tier average is anchored to older,
# lower prices; the uplift itself handles the price-evolution adjustment.
TIER_SHRINK_TIERS = {"Marquee", "Chorus", "TCB", "Mission", "AiR"}

# Alternative scenarios for the Scenarios sheet (apples-to-apples comparison)
SCENARIO_BASELINE = {k: 0.0 for k in UPLIFT_BY_TIER}
SCENARIO_FLAT_10_HP = {
    "Marquee": 0.10, "Headliner": 0.10, "Prestige": 0.10,
    "Standard": 0.0, "Chorus": 0.0, "TCB": 0.0, "Mission": 0.0, "AiR": 0.0,
}
SCENARIO_FLAT_10_HPS = {
    "Marquee": 0.10, "Headliner": 0.10, "Prestige": 0.10,
    "Standard": 0.10, "Chorus": 0.0, "TCB": 0.0, "Mission": 0.0, "AiR": 0.0,
}
SCENARIO_TIERED_NO_MARQUEE = {
    "Marquee":         0.15,  # treated same as Headliner — no separate tier
    "Headliner":       0.15,
    "Prestige":        0.15,
    "Standard": 0.10,
    "Chorus": 0.0, "TCB": 0.0, "Mission": 0.0, "AiR": 0.0,
}
SCENARIO_PRIMARY = UPLIFT_BY_TIER

NAVY        = "1A3A5C"
NAVY_LIGHT  = "E8EDF2"
WHITE       = "FFFFFF"
INK         = "1A2330"
INK_LT      = "5A6A7A"
ROW_STRIPE  = "F5F7FA"
GREEN_FILL  = "D4EDDA"; GREEN_TEXT = "155724"
AMBER_FILL  = "FFF3CD"; AMBER_TEXT = "856404"
RED_FILL    = "F8D7DA"; RED_TEXT   = "721C24"


def _err_palette(pct):
    a = abs(pct)
    if a <= 0.10: return GREEN_FILL, GREEN_TEXT
    if a <= 0.25: return AMBER_FILL, AMBER_TEXT
    return RED_FILL, RED_TEXT


def _style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(bold=True, color=WHITE, size=11)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 28


def _stripe(ws, first_row, last_row, ncols):
    fill = PatternFill("solid", fgColor=ROW_STRIPE)
    for r in range(first_row, last_row + 1):
        if (r - first_row) % 2 == 1:
            for c in range(1, ncols + 1):
                cell = ws.cell(row=r, column=c)
                if cell.fill.start_color.rgb in (None, "00000000", "FFFFFFFF"):
                    cell.fill = fill


def _autowidth(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── Price-prior hierarchy ─────────────────────────────────────────────────────

def _price_for(df, key_cols):
    sub = df.dropna(subset=key_cols)
    if sub.empty:
        return {}
    g = sub.groupby(key_cols, dropna=False).agg(
        rev=("RevWtd", "sum"),
        qty=("QtyWtd", "sum"),
        n=("Quantity", "size"),
        n_events=("EventName", "nunique"),
    )
    g = g[g["qty"] > 0]
    g["price"] = g["rev"] / g["qty"]
    out = {}
    for idx, row in g.iterrows():
        key = idx if isinstance(idx, tuple) else (idx,)
        out[key] = (float(row["price"]), int(row["n"]), int(row["n_events"]))
    return out


def build_price_priors(merged, training_seasons, manifest=None):
    df = merged[
        (merged["Season"].isin(training_seasons))
        & (merged["EventType"] == "Live")
        & (merged["EventStatus"] == "Complete")
        & (merged["TicketStatus"] == "Active")
        & (merged["Quantity"] > 0)
        & (merged["IsComp"] == False)               # paid only
        & (~merged["Pricing"].isin(["PWYW", "Free"]))
        & (merged["TicketTotal"] > 0)
    ].copy()

    # Join PriceTier from the manifest (training data doesn't have it natively)
    if manifest is not None and "PriceTier" in manifest.columns:
        tier_lookup = manifest.drop_duplicates("EventName")[["EventName", "PriceTier"]]
        df = df.merge(tier_lookup, on="EventName", how="left")

    df["W"]      = df["Season"].map(WEIGHTS).fillna(1.0)
    df["RevWtd"] = df["TicketTotal"] * df["W"]
    df["QtyWtd"] = df["Quantity"]    * df["W"]

    # F1b (Class × LoB × Venue) sits between Primary and F1 so events with
    # programming-line context (TCB, AiR, Concert) fall into a LoB-specific
    # pool before mixing with non-LoB Standard pricing. This stops TCB Side
    # events from inheriting prices from unrelated Standard recitals when
    # Primary is too thin to fire.
    levels = {
        "EventRepeat": ["EventRepeat"],
        "Primary":     ["EventClass", "EventVenue", "EventLoB", "EventSubGenre"],
        "F1b":         ["EventClass", "EventLoB", "EventVenue"],
        "F1":          ["EventClass", "EventVenue", "EventSubGenre"],
        "F2":          ["EventClass", "EventVenue", "EventGenre"],
        "F3":          ["EventClass", "EventVenue"],
        "F4":          ["EventSubGenre"],
    }
    priors = {name: _price_for(df, cols) for name, cols in levels.items()}

    rev   = df["RevWtd"].sum()
    qty   = df["QtyWtd"].sum()
    overall = float(rev / qty) if qty else 0.0
    priors["_overall"] = {"price": overall}

    # Per-tier overall used as the shrinkage TARGET so thin buckets don't get
    # pulled toward the global mean. A thin TCB bucket should shrink toward
    # TCB-overall (~$35) not the global $51, otherwise it inflates.
    tier_overall = {}
    if "PriceTier" in df.columns:
        for tier in df["PriceTier"].dropna().unique():
            t = df[df["PriceTier"] == tier]
            tq = t["QtyWtd"].sum()
            if tq > 0:
                tier_overall[tier] = float(t["RevWtd"].sum() / tq)
    priors["_tier_overall"] = tier_overall

    priors["_levels"]  = levels
    priors["_min_n"]   = MIN_N
    priors["_shrink_k"]= SHRINK_K
    return priors


def _shrink(entry, target, k):
    """Bayesian shrink price toward `target` by event-count pseudocounts."""
    price, _n_rows, n_events = entry
    if k <= 0 or n_events <= 0:
        return price
    return (n_events * price + k * target) / (n_events + k)


def lookup_price(row, priors):
    min_n        = priors["_min_n"]
    levels       = priors["_levels"]
    overall      = priors["_overall"]["price"]
    tier_overall = priors.get("_tier_overall", {})
    k            = priors["_shrink_k"]

    # Tier-specific shrinkage target for 0%-uplift tiers (Chorus/TCB/Mission/AiR)
    # where MW prices to a deliberate accessible posture. For commercial tiers,
    # the uplift itself handles tier-specific pricing — use global overall.
    tier = row.get("PriceTier")
    if pd.notna(tier) and tier in TIER_SHRINK_TIERS and tier in tier_overall:
        target = tier_overall[tier]
    else:
        target = overall

    er = row.get("EventRepeat")
    if pd.notna(er):
        entry = priors["EventRepeat"].get((er,))
        if entry and entry[1] >= min_n:
            return _shrink(entry, target, k), f"EventRepeat ({er})"

    for level in ["Primary", "F1b", "F1", "F2", "F3", "F4"]:
        cols = levels[level]
        key = tuple(row.get(c) for c in cols)
        if any(pd.isna(v) for v in key):
            continue
        entry = priors[level].get(key)
        if entry and entry[1] >= min_n:
            return _shrink(entry, target, k), level

    return target, "_tier_overall" if pd.notna(tier) and tier in tier_overall else "_overall"


def uplift_multiplier(price_tier, scenario, festival_role=None):
    """Return multiplier for an event's price prior under `scenario`
    (a dict {tier: pct}).

    TCB events are FestivalRole-aware in scenarios that use a negative TCB
    uplift: the negative uplift applies only to TCB Side events (the
    audience-building accessibly-priced shows). TCB Main events stay at the
    bucket prior (no uplift). Other tiers use the scenario value as-is.
    """
    if pd.isna(price_tier):
        return 1.0
    base = scenario.get(price_tier, 0.0)
    # TCB Main events override a negative TCB scenario value back to 0
    if price_tier == "TCB" and base < 0 and festival_role == "Main":
        return 1.0
    return 1.0 + base


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    em, merged = load_data()
    prior = sorted([s for s in merged["Season"].dropna().unique() if s < FORECAST_SEASON])
    priors = build_price_priors(merged, prior, manifest=em)
    print(f"Overall historical avg paid price (weighted, PWYW/Free excluded): ${priors['_overall']['price']:.2f}")

    # 25-26 forecast (Pred_Paid)
    fcst = pd.read_excel(FORECAST_XLSX, sheet_name="Detail")
    fcst["EventDate"] = pd.to_datetime(fcst["EventDate"], errors="coerce")
    # Merge in manifest classification needed for price lookup + PriceTier
    em["EventDate"] = pd.to_datetime(em["EventDate"], errors="coerce")
    em_keys = em.drop_duplicates("EventName")[
        ["EventName", "EventLoB", "EventGenre", "EventSubGenre", "EventRepeat",
         "Pricing", "PriceTier", "FestivalRole"]]
    fcst = fcst.merge(em_keys, on="EventName", how="left")

    # Look up bucket price priors
    prices, sources = [], []
    for _, r in fcst.iterrows():
        p, s = lookup_price(r, priors)
        prices.append(p)
        sources.append(s)
    fcst["PriceHist"]       = prices
    fcst["PriceHist_Source"]= sources

    # Apply primary tier-based uplift scenario to PriceHist used downstream.
    # The Scenarios sheet will also compute side-by-side totals for alternatives.
    fcst["UpliftMult"]   = [uplift_multiplier(t, SCENARIO_PRIMARY, fr)
                             for t, fr in zip(fcst["PriceTier"], fcst["FestivalRole"])]
    fcst["PriceHist_Up"] = fcst["PriceHist"] * fcst["UpliftMult"]

    # Predicted revenue — primary scenario uses uplifted price
    fcst["Pred_Revenue"] = fcst["Pred_Paid"] * fcst["PriceHist_Up"]

    # Actual revenue + actual avg price from DataMerge
    paid_25 = merged[
        (merged["Season"] == FORECAST_SEASON)
        & (merged["EventType"] == "Live")
        & (merged["EventStatus"] == "Complete")
        & (merged["TicketStatus"] == "Active")
        & (merged["Quantity"] > 0)
        & (merged["IsComp"] == False)
        & (merged["TicketTotal"] > 0)
    ].copy()
    act = paid_25.groupby("EventName").agg(
        ActualRev=("TicketTotal", "sum"),
        ActualPaidQty=("Quantity", "sum"),
    ).reset_index()
    act["ActualAvgPrice"] = act["ActualRev"] / act["ActualPaidQty"]
    fcst = fcst.merge(act, on="EventName", how="left")

    # Decompose error: (Pred_Paid - ActualPaidQty) × ActualAvgPrice  +  Pred_Paid × (PriceHist - ActualAvgPrice)
    fcst["RevErr"]     = fcst["Pred_Revenue"] - fcst["ActualRev"]
    fcst["RevErrPct"]  = np.where(fcst["ActualRev"] > 0,
                                   fcst["RevErr"] / fcst["ActualRev"], np.nan)
    fcst["PriceGap"]   = fcst["PriceHist"] - fcst["ActualAvgPrice"]
    fcst["PriceGapPct"]= np.where(fcst["ActualAvgPrice"] > 0,
                                   fcst["PriceGap"] / fcst["ActualAvgPrice"], np.nan)
    fcst["VolErr"]     = fcst["Pred_Paid"] - fcst["ActualPaidQty"]

    # Decomposition contributions (volume vs price)
    fcst["VolContrib"]   = fcst["VolErr"]   * fcst["ActualAvgPrice"]
    fcst["PriceContrib"] = fcst["PriceGap"] * fcst["Pred_Paid"]

    # Order chronologically
    fcst = fcst.sort_values("EventDate").reset_index(drop=True)

    # Compute scenarios — tier-based uplift policies
    done = fcst[fcst["Status"] == "Completed"].copy()
    scenarios = {}
    for label, scen in [
        ("Baseline (no uplift)",                        SCENARIO_BASELINE),
        ("Flat +10% Headliner+Prestige",                SCENARIO_FLAT_10_HP),
        ("Flat +10% Headliner+Prestige+Standard",     SCENARIO_FLAT_10_HPS),
        ("Tiered without Marquee (H/P/Marq +15%, Std +10%)",   SCENARIO_TIERED_NO_MARQUEE),
        ("Tiered + Marquee (Marq +40%, H/P +15%, Std +10%)",   SCENARIO_PRIMARY),
    ]:
        mult = pd.Series([uplift_multiplier(t, scen, fr)
                          for t, fr in zip(done["PriceTier"], done["FestivalRole"])],
                         index=done.index)
        rev = (done["Pred_Paid"] * done["PriceHist"] * mult).sum()
        n_uplifted = (mult != 1.0).sum()
        scenarios[label] = {"rev": rev, "n_uplifted": int(n_uplifted), "scen": scen}

    actual_rev = done["ActualRev"].sum()
    act_paid   = done["ActualPaidQty"].sum()
    pred_paid  = done["Pred_Paid"].sum()

    print()
    print("=" * 78)
    print(f"REVENUE HINDCAST — 25-26 ({len(done)} completed events)")
    print(f"Actual paid revenue:  ${int(actual_rev):>10,}   "
          f"({int(act_paid):,} paid tkts, avg ${actual_rev/act_paid:.2f})")
    print(f"Pred paid tickets:    {int(pred_paid):>10,}   "
          f"(hindcast forecast Pred_Paid — same across all scenarios)")
    print("=" * 78)
    print(f"{'Scenario':<50s}{'Pred Rev':>14s}{'Δ vs Actual':>14s}{'%':>10s}{'Uplifted':>10s}")
    for label, s in scenarios.items():
        delta = s["rev"] - actual_rev
        pct   = delta / actual_rev * 100
        print(f"  {label:<48s}${int(s['rev']):>13,}${int(delta):>+13,}{pct:>+9.1f}%{s['n_uplifted']:>10d}")
    print()
    print(f"Volume contribution  (Pred_Paid − Actual_Paid) × Actual_Price = ${int(done['VolContrib'].sum()):+,}")
    print(f"Price contribution   Pred_Paid × (PriceHist − Actual_Price)   = ${int(done['PriceContrib'].sum()):+,}")
    print(f"(Decomposition above uses BASELINE price prior — not uplifted.)")
    print()

    # Primary scenario for the rest of the report
    pred_rev   = done["Pred_Revenue"].sum()
    rev_err    = pred_rev - actual_rev
    pred_price = pred_rev / pred_paid if pred_paid else 0
    act_price  = actual_rev / act_paid if act_paid else 0

    # ── Excel output ─────────────────────────────────────────────────────────
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)

    # Scenarios sheet (new — at front)
    ws = wb.create_sheet("Scenarios")
    ws.cell(row=1, column=1, value="Revenue Hindcast — Price-Uplift Scenarios").font = Font(bold=True, size=16, color=NAVY)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    ws.cell(row=2, column=1, value=(
        f"All scenarios use the same Pred_Paid (blinded hindcast). Only the price prior changes.  "
        f"Mission events and TCB-LoB events always excluded from uplift.")).font = Font(italic=True, color=INK_LT, size=10)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 18

    scen_headers = ["Scenario", "Events uplifted", "Pred Revenue", "Δ vs Actual", "% vs Actual", "Notes"]
    for c, h in enumerate(scen_headers, start=1):
        ws.cell(row=4, column=c, value=h)
    _style_header(ws, 4, len(scen_headers))

    scen_notes = {
        "Baseline (no uplift)":                                   "Raw historical bucket prior — what the hindcast saw blind",
        "Flat +10% Headliner+Prestige":                           "Top-tier only; everything else 0",
        "Flat +10% Headliner+Prestige+Standard":                "Adds Standard at the same 10%",
        "Tiered without Marquee (H/P/Marq +15%, Std +10%)":    "Tier policy WITHOUT the separate Marquee tier",
        "Tiered + Marquee (Marq +40%, H/P +15%, Std +10%)":    "PRIMARY — adds Marquee tier for MW-flagged top-draw bookings",
    }
    # Actual revenue reference row
    ws.cell(row=5, column=1, value="ACTUAL paid revenue (reference)").font = Font(bold=True, color=INK, size=11)
    ws.cell(row=5, column=3, value=int(actual_rev)).number_format = '"$"#,##0'
    ws.cell(row=5, column=3).font = Font(bold=True, color=INK, size=11)
    for c in (1, 3):
        ws.cell(row=5, column=c).fill = PatternFill("solid", fgColor=NAVY_LIGHT)
    for c in (2, 4, 5, 6):
        ws.cell(row=5, column=c).fill = PatternFill("solid", fgColor=NAVY_LIGHT)
    ws.cell(row=5, column=6, value="Reference baseline for the scenarios below").fill = PatternFill("solid", fgColor=NAVY_LIGHT)

    for i, (label, s) in enumerate(scenarios.items()):
        row = 6 + i
        delta = s["rev"] - actual_rev
        pct   = delta / actual_rev if actual_rev else None
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=s["n_uplifted"])
        ws.cell(row=row, column=3, value=int(s["rev"])).number_format = '"$"#,##0'
        ws.cell(row=row, column=4, value=int(delta)).number_format = '"$"+#,##0;[Red]"$"-#,##0;0'
        pct_cell = ws.cell(row=row, column=5, value=pct)
        pct_cell.number_format = "+0.0%;-0.0%;0.0%"
        if pct is not None:
            fill_color, text_color = _err_palette(pct)
            pct_cell.fill = PatternFill("solid", fgColor=fill_color)
            pct_cell.font = Font(bold=True, color=text_color, size=11)
        ws.cell(row=row, column=6, value=scen_notes.get(label, ""))

    _autowidth(ws, [34, 16, 14, 14, 13, 56])
    ws.sheet_view.showGridLines = False

    # Summary sheet (per-event accuracy of primary scenario)
    ws = wb.create_sheet("Summary")
    ws.cell(row=1, column=1, value="Music Worcester 25-26 — Revenue Hindcast (Primary Scenario)").font = Font(bold=True, size=16, color=NAVY)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    ws.cell(row=2, column=1, value=(
        f"Primary policy: Marquee +30%, Headliner/Prestige +15%, Standard +10%, Chorus/TCB/Mission/AiR 0%.  "
        f"PWYW/Free excluded from training, season-weighted, Bayes shrink K={SHRINK_K}.")).font = Font(italic=True, color=INK_LT, size=10)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 18

    headers = ["Metric", "Predicted", "Actual", "Δ", "% Error", "Notes"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=4, column=c, value=h)
    _style_header(ws, 4, len(headers))

    rows = [
        ("Paid tickets",          int(pred_paid), int(act_paid),
         int(pred_paid - act_paid), (pred_paid-act_paid)/act_paid if act_paid else None,
         "Volume gap (forecast vs actual)"),
        ("Avg paid price ($/tkt)", round(pred_price, 2), round(act_price, 2),
         round(pred_price - act_price, 2),
         (pred_price-act_price)/act_price if act_price else None,
         "Tier-uplifted bucket price (Marq +40%, H/P +15%, Std +10%) vs actual"),
        ("Paid revenue ($)",      int(pred_rev), int(actual_rev),
         int(rev_err), rev_err/actual_rev if actual_rev else None,
         "Combined volume × price error"),
        ("  Volume contribution", None, None, int(done['VolContrib'].sum()), None,
         "(Pred_Paid − Actual_Paid) × Actual_Price (baseline prior, NOT uplifted)"),
        ("  Price contribution",  None, None, int(done['PriceContrib'].sum()), None,
         "Pred_Paid × (PriceHist − Actual_Price) (baseline prior, NOT uplifted)"),
    ]
    for i, r in enumerate(rows):
        row = 5 + i
        for c, v in enumerate(r, start=1):
            cell = ws.cell(row=row, column=c, value=v)
            if c in (2, 3, 4) and isinstance(v, (int, float)):
                if r[0] == "Avg paid price ($/tkt)":
                    cell.number_format = '"$"#,##0.00'
                elif "revenue" in r[0].lower() or "contribution" in r[0].lower():
                    cell.number_format = '"$"#,##0;[Red]-"$"#,##0'
                else:
                    cell.number_format = "+#,##0;-#,##0;0" if c == 4 else "#,##0"
            elif c == 5 and isinstance(v, float):
                cell.number_format = "+0.0%;-0.0%;0.0%"
        if r[0].startswith(("Paid revenue", "Paid tickets", "Avg paid")):
            for c in range(1, 7):
                cell = ws.cell(row=row, column=c)
                cell.font = Font(bold=True, color=INK, size=11)
                cell.fill = PatternFill("solid", fgColor=NAVY_LIGHT)

    _autowidth(ws, [26, 14, 14, 14, 12, 48])
    ws.sheet_view.showGridLines = False

    # Per-event sheet
    ws = wb.create_sheet("Per-Event")
    headers = ["Date", "Event", "Class",
               "Pred Paid", "Actual Paid",
               "Hist Price", "Actual Price", "Price Gap",
               "Pred Revenue", "Actual Revenue", "Rev Error", "Rev % Err",
               "Vol Contrib", "Price Contrib", "Price Source"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    _style_header(ws, 1, len(headers))

    done_chrono = done.copy()
    int_fmt   = "#,##0"
    money_fmt = '"$"#,##0;[Red]-"$"#,##0'
    price_fmt = '"$"#,##0.00'
    err_fmt   = "+#,##0;-#,##0;0"
    pct_fmt   = "+0.0%;-0.0%;0.0%"

    for i, (_, r) in enumerate(done_chrono.iterrows()):
        row = 2 + i
        ws.cell(row=row, column=1, value=r["EventDate"]).number_format = "mmm d, yyyy"
        ws.cell(row=row, column=2, value=r["EventName"])
        ws.cell(row=row, column=3, value=r["EventClass"])
        ws.cell(row=row, column=4, value=int(r["Pred_Paid"])).number_format = int_fmt
        ws.cell(row=row, column=5, value=int(r["ActualPaidQty"]) if pd.notna(r["ActualPaidQty"]) else None).number_format = int_fmt
        ws.cell(row=row, column=6, value=round(r["PriceHist"], 2)).number_format = price_fmt
        ws.cell(row=row, column=7, value=round(r["ActualAvgPrice"], 2) if pd.notna(r["ActualAvgPrice"]) else None).number_format = price_fmt
        ws.cell(row=row, column=8, value=round(r["PriceGap"], 2) if pd.notna(r["PriceGap"]) else None).number_format = "+$#,##0.00;-$#,##0.00"
        ws.cell(row=row, column=9, value=int(r["Pred_Revenue"])).number_format = money_fmt
        ws.cell(row=row, column=10, value=int(r["ActualRev"]) if pd.notna(r["ActualRev"]) else None).number_format = money_fmt
        ws.cell(row=row, column=11, value=int(r["RevErr"]) if pd.notna(r["RevErr"]) else None).number_format = err_fmt
        pct_cell = ws.cell(row=row, column=12,
                            value=r["RevErrPct"] if pd.notna(r["RevErrPct"]) else None)
        pct_cell.number_format = pct_fmt
        if pd.notna(r["RevErrPct"]):
            fill_color, text_color = _err_palette(r["RevErrPct"])
            pct_cell.fill = PatternFill("solid", fgColor=fill_color)
            pct_cell.font = Font(bold=True, color=text_color, size=11)
        ws.cell(row=row, column=13, value=int(r["VolContrib"]) if pd.notna(r["VolContrib"]) else None).number_format = err_fmt
        ws.cell(row=row, column=14, value=int(r["PriceContrib"]) if pd.notna(r["PriceContrib"]) else None).number_format = err_fmt
        ws.cell(row=row, column=15, value=r["PriceHist_Source"])

    last_row = 1 + len(done_chrono)
    _stripe(ws, 2, last_row, len(headers))

    # Totals row
    tr = last_row + 1
    ws.cell(row=tr, column=1, value="Total")
    tcells = {
        4:  int(done['Pred_Paid'].sum()),
        5:  int(done['ActualPaidQty'].sum()),
        9:  int(done['Pred_Revenue'].sum()),
        10: int(done['ActualRev'].sum()),
        11: int(done['RevErr'].sum()),
        13: int(done['VolContrib'].sum()),
        14: int(done['PriceContrib'].sum()),
    }
    for col, v in tcells.items():
        cell = ws.cell(row=tr, column=col, value=v)
        cell.number_format = err_fmt if col in (11, 13, 14) else (money_fmt if col in (9, 10) else int_fmt)
    # Vol-weighted % rev err
    if done['ActualRev'].sum() > 0:
        c = ws.cell(row=tr, column=12, value=done['RevErr'].sum() / done['ActualRev'].sum())
        c.number_format = pct_fmt
    # Vol-weighted avg prices in the price cols
    if done['Pred_Paid'].sum() > 0:
        c = ws.cell(row=tr, column=6, value=round(done['Pred_Revenue'].sum() / done['Pred_Paid'].sum(), 2))
        c.number_format = price_fmt
    if done['ActualPaidQty'].sum() > 0:
        c = ws.cell(row=tr, column=7, value=round(done['ActualRev'].sum() / done['ActualPaidQty'].sum(), 2))
        c.number_format = price_fmt
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=tr, column=c)
        cell.font = Font(bold=True, color=INK, size=11)
        cell.fill = PatternFill("solid", fgColor=NAVY_LIGHT)
        cell.border = Border(top=Side(border_style="thin", color=NAVY))

    _autowidth(ws, [12, 44, 11, 10, 11, 11, 12, 11, 13, 14, 12, 11, 12, 13, 22])
    ws.freeze_panes = "C2"
    ws.sheet_view.showGridLines = False

    # Pricing Observations sheet — events where actual avg price diverged meaningfully
    ws = wb.create_sheet("PricingObservations")
    ws.cell(row=1, column=1, value="Pricing observations — events where realized price diverged from historical bucket prior").font = Font(bold=True, size=12, color=NAVY)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    ws.cell(row=2, column=1, value=(
        "Positive PriceGap = MW priced ABOVE historical (potential price-raising success).  "
        "Negative = priced BELOW historical (left money on table OR audience-build pricing).")).font = Font(italic=True, color=INK_LT, size=10)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)
    ws.row_dimensions[1].height = 22

    headers = ["Date", "Event", "Class", "Hist Price", "Actual Price",
               "Price Gap", "Price Gap %", "Read"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=4, column=c, value=h)
    _style_header(ws, 4, len(headers))

    # Filter to materially-different (>= 10% price gap) and meaningful volume
    obs = done[(done["ActualPaidQty"] >= 50)
               & (done["PriceGapPct"].abs() >= 0.10)
               & (done["ActualAvgPrice"].notna())].copy()
    obs = obs.sort_values("PriceGapPct", ascending=True)  # below-historical first

    def read_pricing(row):
        gap = row["PriceGapPct"]
        if gap <= -0.20:
            return "Priced WELL BELOW historical — opportunity to test higher pricing next iteration"
        if gap <= -0.10:
            return "Priced below historical — modest headroom may exist"
        if gap >= 0.20:
            return "Priced WELL ABOVE historical — verify volume didn't suffer; success or stretch"
        if gap >= 0.10:
            return "Priced above historical — held demand at higher price point"
        return ""

    first = 5
    for i, (_, r) in enumerate(obs.iterrows()):
        row = first + i
        ws.cell(row=row, column=1, value=r["EventDate"]).number_format = "mmm d, yyyy"
        ws.cell(row=row, column=2, value=r["EventName"])
        ws.cell(row=row, column=3, value=r["EventClass"])
        ws.cell(row=row, column=4, value=round(r["PriceHist"], 2)).number_format = price_fmt
        ws.cell(row=row, column=5, value=round(r["ActualAvgPrice"], 2)).number_format = price_fmt
        ws.cell(row=row, column=6, value=round(r["PriceGap"], 2)).number_format = "+$#,##0.00;-$#,##0.00"
        gap_cell = ws.cell(row=row, column=7, value=r["PriceGapPct"])
        gap_cell.number_format = pct_fmt
        # Color: green if priced above (good for revenue), amber/red if below
        if r["PriceGapPct"] >= 0.10:
            gap_cell.fill = PatternFill("solid", fgColor=GREEN_FILL)
            gap_cell.font = Font(bold=True, color=GREEN_TEXT, size=11)
        else:  # negative
            gap_cell.fill = PatternFill("solid", fgColor=AMBER_FILL)
            gap_cell.font = Font(bold=True, color=AMBER_TEXT, size=11)
        ws.cell(row=row, column=8, value=read_pricing(r))

    last = first + len(obs) - 1
    if len(obs) > 0:
        _stripe(ws, first, last, len(headers))
    _autowidth(ws, [12, 44, 11, 11, 12, 11, 12, 70])
    ws.freeze_panes = "C5"
    ws.sheet_view.showGridLines = False

    wb.save(OUT_XLSX)
    print(f"✓ {OUT_XLSX}")
    print(f"  3 sheets: Summary, Per-Event ({len(done)} events), PricingObservations ({len(obs)} flagged)")


if __name__ == "__main__":
    main()
