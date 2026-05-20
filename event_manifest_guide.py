"""
Human-readable event catalog — aggregates EventManifest.xlsx to EventId
level and writes a finance-friendly workbook with a glossary.

Output: Event_Catalog.xlsx at repo root (one row per event).
"""
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

os.chdir("/Users/antho/Documents/WPI-MW")

MANIFEST = "EventManifest.xlsx"
OUT      = "Event_Catalog.xlsx"

NAVY        = "1A3A5C"
NAVY_LIGHT  = "E8EDF2"
WHITE       = "FFFFFF"
INK         = "1A2330"
INK_LT      = "5A6A7A"
ROW_STRIPE  = "F5F7FA"

# Expand internal codes to friendly labels for the catalog.
LOB_LABEL = {
    "Concert": "Regular Concert",
    "TCB":     "The Complete Bach festival",
    "AiR":     "Artist in Residence",
    "Bach":    "Bach Programming (legacy tag)",
}
CLASS_LABEL = {
    "Headliner": "Headliner (top draw, broad audience)",
    "Prestige":  "Prestige (world-class specialist)",
    "Standard":  "Standard (regular programming)",
    "Mission":   "Mission (community / education focus)",
}
PRICETIER_LABEL = {
    "Marquee":         "Marquee (top-draw bookings, priced premium)",
    "Headliner":       "Headliner (top tier)",
    "Prestige":        "Prestige (specialist tier)",
    "Standard":        "Standard (typical commercial concert programming)",
    "Chorus":          "Chorus (accessibly priced)",
    "TCB":             "TCB (festival pricing posture)",
    "Mission":         "Mission (community pricing)",
    "AiR":             "AiR (community pricing)",
}
PRICING_LABEL = {
    "PWYW": "Pay What You Wish",
    "Free": "Free / fully comped",
}


def _style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(bold=True, color=WHITE, size=11)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 28


def _stripe(ws, first_row, last_row, ncols):
    fill = PatternFill("solid", fgColor=ROW_STRIPE)
    for r in range(first_row, last_row + 1):
        if (r - first_row) % 2 == 1:
            for c in range(1, ncols + 1):
                cell = ws.cell(row=r, column=c)
                if cell.fill.start_color.rgb in (None, "00000000", "FFFFFFFF"):
                    cell.fill = fill


def _autowidth(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_glossary(ws):
    ws.cell(row=1, column=1, value="Event Catalog — Glossary").font = Font(bold=True, size=15, color=NAVY)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    ws.cell(row=2, column=1, value="A finance-friendly view of every event MW programs. One row per event. The Instances count notes multi-night runs or events sold through multiple ticketing channels.").font = Font(italic=True, color=INK_LT, size=10)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=3)
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 18

    # Column definitions table
    ws.cell(row=4, column=1, value="Column definitions").font = Font(bold=True, color=NAVY, size=11)

    col_defs = [
        ("Season",          "MW programming season (e.g. 25-26 = 2025-2026)"),
        ("Date",            "Event date (first show if multi-night)"),
        ("Event",           "Event name as it appears on tickets"),
        ("Venue",           "Where the event takes place"),
        ("Capacity",        "Room capacity (paid + comp seats)"),
        ("Instances",       "Number of Salesforce ticketing instances for this event. Usually = number of performances (e.g. a 2-night run), but can also count separate sales channels for the same show (e.g. main + ETO/External Ticket Office, or main + Subscription Sales)."),
        ("Status",          "Complete = happened; Future = upcoming; Canceled / Postponed"),
        ("Programming Line", "What part of MW's programming this is — Regular Concert, The Complete Bach festival, Artist in Residence"),
        ("Class",           "Booking tier — Headliner, Prestige, Standard, Mission"),
        ("Genre",           "Musical category (Classical / Jazz / Dance / Choral / etc.)"),
        ("Sub-genre",       "Finer grain (e.g. Piano / Cantata / Modern Dance)"),
        ("Pricing Tier",    "How this event is priced — see tier table below"),
        ("Festival Role",   "Main (anchor slot) or Side (side event); blank for non-festival"),
        ("Recurring Series", "Tag if part of an annual series (e.g. DTH, Messiah, Cantatathon)"),
        ("Pricing Note",    "PWYW = Pay What You Wish; Free = fully comped; blank for standard pricing"),
    ]
    for c, h in enumerate(["Column", "What it means"], start=1):
        ws.cell(row=6, column=c, value=h)
    _style_header(ws, 6, 2)
    for i, (col, desc) in enumerate(col_defs):
        row = 7 + i
        ws.cell(row=row, column=1, value=col).font = Font(bold=True, color=NAVY, size=10)
        ws.cell(row=row, column=2, value=desc).font = Font(size=10, color=INK)
    last = 7 + len(col_defs) - 1
    _stripe(ws, 7, last, 2)

    # Pricing tier reference
    block = last + 3
    ws.cell(row=block, column=1, value="Pricing tiers").font = Font(bold=True, color=NAVY, size=11)
    for c, h in enumerate(["Tier", "Description"], start=1):
        ws.cell(row=block + 2, column=c, value=h)
    _style_header(ws, block + 2, 2)
    for i, (tier, desc) in enumerate(PRICETIER_LABEL.items()):
        row = block + 3 + i
        ws.cell(row=row, column=1, value=tier).font = Font(bold=True, color=NAVY, size=10)
        ws.cell(row=row, column=2, value=desc).font = Font(size=10, color=INK)
    _stripe(ws, block + 3, block + 2 + len(PRICETIER_LABEL), 2)

    # Programming line reference
    block2 = block + 3 + len(PRICETIER_LABEL) + 2
    ws.cell(row=block2, column=1, value="Programming lines (EventLoB)").font = Font(bold=True, color=NAVY, size=11)
    for c, h in enumerate(["Code", "Description"], start=1):
        ws.cell(row=block2 + 2, column=c, value=h)
    _style_header(ws, block2 + 2, 2)
    for i, (k, v) in enumerate(LOB_LABEL.items()):
        row = block2 + 3 + i
        ws.cell(row=row, column=1, value=k).font = Font(bold=True, color=NAVY, size=10)
        ws.cell(row=row, column=2, value=v).font = Font(size=10, color=INK)
    _stripe(ws, block2 + 3, block2 + 2 + len(LOB_LABEL), 2)

    # Class reference
    block3 = block2 + 3 + len(LOB_LABEL) + 2
    ws.cell(row=block3, column=1, value="Event classes").font = Font(bold=True, color=NAVY, size=11)
    for c, h in enumerate(["Class", "Description"], start=1):
        ws.cell(row=block3 + 2, column=c, value=h)
    _style_header(ws, block3 + 2, 2)
    for i, (k, v) in enumerate(CLASS_LABEL.items()):
        row = block3 + 3 + i
        ws.cell(row=row, column=1, value=k).font = Font(bold=True, color=NAVY, size=10)
        ws.cell(row=row, column=2, value=v).font = Font(size=10, color=INK)
    _stripe(ws, block3 + 3, block3 + 2 + len(CLASS_LABEL), 2)

    _autowidth(ws, [22, 90])
    ws.sheet_view.showGridLines = False


def build_events(ws, df):
    headers = ["Season", "Date", "Event", "Venue", "Capacity", "Instances",
               "Status", "Programming Line", "Class", "Genre", "Sub-genre",
               "Pricing Tier", "Festival Role", "Recurring Series", "Pricing Note"]
    ncols = len(headers)
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    _style_header(ws, 1, ncols)

    def safe(v):
        return "" if pd.isna(v) else v

    def lob_text(v):
        if pd.isna(v):
            return ""
        return LOB_LABEL.get(v, v)

    def pricing_text(v):
        if pd.isna(v):
            return ""
        return PRICING_LABEL.get(v, v)

    for i, (_, r) in enumerate(df.iterrows()):
        row = 2 + i
        ws.cell(row=row, column=1,  value=safe(r["Season"]))
        d = r["EventDate"]
        c2 = ws.cell(row=row, column=2, value=d if pd.notna(d) else None)
        c2.number_format = "mmm d, yyyy"
        ws.cell(row=row, column=3,  value=safe(r["EventName"]))
        ws.cell(row=row, column=4,  value=safe(r["EventVenue"]))
        cap = r["EventCapacity"]
        c5 = ws.cell(row=row, column=5, value=int(cap) if pd.notna(cap) else None)
        c5.number_format = "#,##0"
        ws.cell(row=row, column=6,  value=int(r["Instances"])).number_format = "#,##0"
        ws.cell(row=row, column=7,  value=safe(r["EventStatus"]))
        ws.cell(row=row, column=8,  value=lob_text(r["EventLoB"]))
        ws.cell(row=row, column=9,  value=safe(r["EventClass"]))
        ws.cell(row=row, column=10, value=safe(r["EventGenre"]))
        ws.cell(row=row, column=11, value=safe(r["EventSubGenre"]))
        ws.cell(row=row, column=12, value=safe(r["PriceTier"]))
        ws.cell(row=row, column=13, value=safe(r["FestivalRole"]))
        ws.cell(row=row, column=14, value=safe(r["EventRepeat"]))
        ws.cell(row=row, column=15, value=pricing_text(r["Pricing"]))

    last_row = 1 + len(df)
    _stripe(ws, 2, last_row, ncols)

    # Auto-filter on header row
    ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{last_row}"
    ws.freeze_panes = "D2"
    _autowidth(ws, [8, 13, 44, 22, 10, 7, 12, 28, 12, 14, 18, 17, 13, 18, 22])
    ws.sheet_view.showGridLines = False


def main():
    m = pd.read_excel(MANIFEST)

    # Drop test events; restrict to actual performances (skip subscription bundles + placeholders)
    m = m[~m["EventName"].fillna("").str.contains("test|TEST|Test|Placeholder", case=False, regex=True)]
    m = m[m["EventType"].isin(["Live", "Virtual", "Pandemic"])].copy()
    m["EventDate"] = pd.to_datetime(m["EventDate"], errors="coerce")

    # Aggregate to EventId: take first row's metadata, count instances, use first date
    agg = (
        m.sort_values("EventDate")
         .groupby("EventId", dropna=False)
         .agg(
             Instances     = ("InstanceId", "count"),
             EventDate     = ("EventDate", "first"),
             EventName     = ("EventName", "first"),
             EventVenue    = ("EventVenue", "first"),
             EventCapacity = ("EventCapacity", "first"),
             Season        = ("Season", "first"),
             EventStatus   = ("EventStatus", "first"),
             EventLoB      = ("EventLoB", "first"),
             EventClass    = ("EventClass", "first"),
             EventGenre    = ("EventGenre", "first"),
             EventSubGenre = ("EventSubGenre", "first"),
             EventRepeat   = ("EventRepeat", "first"),
             PriceTier     = ("PriceTier", "first"),
             FestivalRole  = ("FestivalRole", "first"),
             Pricing       = ("Pricing", "first"),
         )
         .reset_index()
    )

    # Sort: season desc (most recent first), then date asc within season
    agg = agg.sort_values(["Season", "EventDate"], ascending=[False, True]).reset_index(drop=True)

    wb = Workbook()
    wb.remove(wb.active)
    build_events(wb.create_sheet("Events"), agg)
    build_glossary(wb.create_sheet("Glossary"))
    wb.save(OUT)

    seasons = sorted(agg["Season"].dropna().unique())
    print(f"  ✓ {OUT}")
    print(f"  {len(agg)} events  ·  spanning {seasons[0]} to {seasons[-1]}")
    print(f"  PriceTier distribution: {dict(agg['PriceTier'].value_counts(dropna=False))}")


if __name__ == "__main__":
    main()
