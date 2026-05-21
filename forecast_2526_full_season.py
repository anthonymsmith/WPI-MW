"""
Full 25-26 season forecast — one clean tab.

Includes every 25-26 live event from the manifest (completed + upcoming),
with the final adjusted prediction and 90% CI. Meant for planning, budgeting,
and sharing — no diagnostic columns.

Output: forecasting/Forecast_2526_FullSeason.xlsx with three sheets:
  - "Summary"             — paid+comp / paid / comp totals + accuracy
  - "Forecast vs Actuals" — lean per-event business view with color-coded % Error
  - "Detail"              — full technical table (all CI bounds + comp-rate source)
Detail sheet columns: EventDate, EventName, EventVenue, EventClass,
         EventCapacity, Status, Actual, Pred_Adj, Pred_Adj_Lo, Pred_Adj_Hi
"""
import os
from datetime import date
import pandas as pd
import numpy as np
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell

os.chdir("/Users/antho/Documents/WPI-MW")

from forecast_2526_comparison import (
    load_data, get_training_df, build_hierarchy_models,
    predict_model_a, cap_at_capacity, build_pwyw_lift, build_side_lift,
    INCLUDE_COMPS,
)
from forecast_artist_adjustment import apply_artist_adjustment
from forecast_bucket_ci import apply_bucket_ci
from forecast_comp_split import build_comp_rates, apply_comp_split

FORECAST_SEASON = "25-26"
OUT_XLSX = "forecasting/Forecast_2526_FullSeason.xlsx"

NAVY        = "1A3A5C"
NAVY_LIGHT  = "E8EDF2"
WHITE       = "FFFFFF"
INK         = "1A2330"
INK_LT      = "5A6A7A"
ROW_STRIPE  = "F5F7FA"
GREEN_FILL  = "D4EDDA"; GREEN_TEXT = "155724"
AMBER_FILL  = "FFF3CD"; AMBER_TEXT = "856404"
RED_FILL    = "F8D7DA"; RED_TEXT   = "721C24"


def _err_palette(pct):
    a = abs(pct)
    if a <= 0.10:  return GREEN_FILL, GREEN_TEXT
    if a <= 0.25:  return AMBER_FILL, AMBER_TEXT
    return RED_FILL, RED_TEXT


def _style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(bold=True, color=WHITE, size=11)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 28


def _stripe(ws, first_row, last_row, ncols):
    fill = PatternFill("solid", fgColor=ROW_STRIPE)
    for r in range(first_row, last_row + 1):
        if (r - first_row) % 2 == 1:
            for c in range(1, ncols + 1):
                cell = ws.cell(row=r, column=c)
                if cell.fill.start_color.rgb in (None, "00000000", "FFFFFFFF"):
                    cell.fill = fill


def _autowidth(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_summary_sheet(ws, summary_df, n_total, actual_total, pred_total):
    """3-row summary with a headline title above."""
    today = f"{date.today():%B %-d, %Y}"
    ws.cell(row=1, column=1, value="Music Worcester 25-26 Season Forecast Performance").font = \
        Font(bold=True, size=16, color=NAVY)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    pct = (pred_total - actual_total) / actual_total * 100 if actual_total else 0
    subtitle = (f"Pre-season forecast across {n_total} completed events  ·  "
                f"Net forecast {pred_total:,} vs. actual {actual_total:,}  "
                f"({pct:+.2f}%)  ·  Refreshed {today}")
    ws.cell(row=2, column=1, value=subtitle).font = Font(italic=True, color=INK_LT, size=10)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 18

    # Headers at row 4
    headers = ["Metric", "Events", "Actual", "Predicted", "Net Error",
               "% Error", "WAPE", "Avg Per-Event Bias"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=4, column=c, value=h)
    _style_header(ws, 4, len(headers))

    # Data rows 5-7
    for i, r in summary_df.iterrows():
        row = 5 + i
        ws.cell(row=row, column=1, value=r["Metric"])
        ws.cell(row=row, column=2, value=int(r["n"]))
        ws.cell(row=row, column=3, value=int(r["Actual"])).number_format = "#,##0"
        ws.cell(row=row, column=4, value=int(r["Predicted"])).number_format = "#,##0"
        ws.cell(row=row, column=5, value=int(r["Net Error"])).number_format = "+#,##0;-#,##0;0"
        ws.cell(row=row, column=6, value=r["% Error"] / 100).number_format = "+0.0%;-0.0%;0.0%"
        ws.cell(row=row, column=7, value=r["WAPE %"] / 100).number_format = "0.0%"
        ws.cell(row=row, column=8, value=r["Bias %"] / 100).number_format = "+0.0%;-0.0%;0.0%"
        # Bold the Total row
        if r["Metric"].startswith("Total"):
            for c in range(1, 9):
                cell = ws.cell(row=row, column=c)
                cell.font = Font(bold=True, color=INK, size=11)
                cell.fill = PatternFill("solid", fgColor=NAVY_LIGHT)

    # Footer legend
    legend_row = 9
    ws.cell(row=legend_row, column=1, value=(
        "WAPE = total absolute error / total actual (volume-weighted).  "
        "Bias = average per-event signed error (>0 = over-forecast)."
    )).font = Font(italic=True, color=INK_LT, size=9)
    ws.merge_cells(start_row=legend_row, start_column=1, end_row=legend_row, end_column=8)

    _autowidth(ws, [22, 9, 12, 12, 12, 11, 10, 19])
    ws.sheet_view.showGridLines = False


def write_business_sheet(ws, out):
    """Forecast vs Actuals — lean business view."""
    df = out.copy()
    df["Error"]      = df["Pred_Adj"]  - df["Actual"]
    df["ErrPct"]     = np.where(df["Actual"] > 0,      df["Error"]     / df["Actual"],      np.nan)
    df["PaidErr"]    = df["Pred_Paid"] - df["Actual_Paid"]
    df["PaidErrPct"] = np.where(df["Actual_Paid"] > 0, df["PaidErr"]   / df["Actual_Paid"], np.nan)
    df["CompErr"]    = df["Pred_Comp"] - df["Actual_Comp"]
    df["CompErrPct"] = np.where(df["Actual_Comp"] > 0, df["CompErr"]   / df["Actual_Comp"], np.nan)

    headers = ["Date", "Event", "Venue", "Class", "Capacity", "Status",
               "Forecast (Total)", "Actual (Total)", "Error", "% Error",
               "Forecast Paid", "Actual Paid", "Paid Error", "Paid % Error",
               "Forecast Comp", "Actual Comp", "Comp Error", "Comp % Error"]
    ncols = len(headers)

    ws.cell(row=1, column=1, value="Forecast vs. Actual — 2025–26 Season").font = \
        Font(bold=True, size=14, color=NAVY)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.row_dimensions[1].height = 22

    for c, h in enumerate(headers, start=1):
        ws.cell(row=3, column=c, value=h)
    _style_header(ws, 3, ncols)

    int_fmt    = "#,##0"
    err_fmt    = "+#,##0;-#,##0;0"
    errpct_fmt = "+0.0%;-0.0%;0.0%"

    def _put_int(row, col, v, fmt=int_fmt):
        cell = ws.cell(row=row, column=col, value=int(v) if pd.notna(v) else None)
        cell.number_format = fmt
        return cell

    def _put_errpct(row, col, v):
        cell = ws.cell(row=row, column=col, value=v if pd.notna(v) else None)
        cell.number_format = errpct_fmt
        if pd.notna(v):
            fill_color, text_color = _err_palette(v)
            cell.fill = PatternFill("solid", fgColor=fill_color)
            cell.font = Font(bold=True, color=text_color, size=11)
        return cell

    first_data_row = 4
    for i, (_, r) in enumerate(df.iterrows()):
        row = first_data_row + i
        ws.cell(row=row, column=1, value=r["EventDate"]).number_format = 'mmm d, yyyy'
        ws.cell(row=row, column=2, value=r["EventName"])
        ws.cell(row=row, column=3, value=r["EventVenue"])
        ws.cell(row=row, column=4, value=r["EventClass"])
        _put_int(row, 5,  r["EventCapacity"])
        ws.cell(row=row, column=6, value=r["Status"])
        _put_int(row, 7,  r["Pred_Adj"])
        _put_int(row, 8,  r["Actual"])
        _put_int(row, 9,  r["Error"],   err_fmt)
        _put_errpct(row, 10, r["ErrPct"])
        _put_int(row, 11, r["Pred_Paid"])
        _put_int(row, 12, r["Actual_Paid"])
        _put_int(row, 13, r["PaidErr"], err_fmt)
        _put_errpct(row, 14, r["PaidErrPct"])
        _put_int(row, 15, r["Pred_Comp"])
        _put_int(row, 16, r["Actual_Comp"])
        _put_int(row, 17, r["CompErr"], err_fmt)
        _put_errpct(row, 18, r["CompErrPct"])

    last_row = first_data_row + len(df) - 1
    _stripe(ws, first_data_row, last_row, ncols)

    # Totals row
    tr = last_row + 1
    ws.cell(row=tr, column=1, value="Total")
    completed = df[df["Status"] == "Completed"]
    int_cols = {
        7:  completed["Pred_Adj"].sum(),    8:  completed["Actual"].sum(),
        9:  completed["Error"].sum(),
        11: completed["Pred_Paid"].sum(),   12: completed["Actual_Paid"].sum(),
        13: completed["PaidErr"].sum(),
        15: completed["Pred_Comp"].sum(),   16: completed["Actual_Comp"].sum(),
        17: completed["CompErr"].sum(),
    }
    for col, v in int_cols.items():
        c = ws.cell(row=tr, column=col, value=int(v))
        c.number_format = err_fmt if col in (9, 13, 17) else int_fmt
    # Volume-weighted % error for each block
    def _vol_pct(err_sum, actual_sum):
        return err_sum / actual_sum if actual_sum else None
    for col, err_col, act_col in [
        (10, "Error",   "Actual"),
        (14, "PaidErr", "Actual_Paid"),
        (18, "CompErr", "Actual_Comp"),
    ]:
        pct = _vol_pct(completed[err_col].sum(), completed[act_col].sum())
        if pct is not None:
            c = ws.cell(row=tr, column=col, value=pct)
            c.number_format = errpct_fmt
    for c in range(1, ncols + 1):
        cell = ws.cell(row=tr, column=c)
        cell.font = Font(bold=True, color=INK, size=11)
        cell.fill = PatternFill("solid", fgColor=NAVY_LIGHT)
        cell.border = Border(top=Side(border_style="thin", color=NAVY))

    # Column widths
    _autowidth(ws, [12, 44, 22, 11, 10, 12,
                    14, 13, 10, 10,
                    14, 12, 11, 12,
                    14, 12, 11, 12])
    ws.freeze_panes = "C4"
    ws.sheet_view.showGridLines = False


def write_detail_sheet(ws, out):
    """Full technical detail — all columns including CI bounds and comp-rate source."""
    cols = ["EventDate", "EventName", "EventVenue", "EventClass", "EventCapacity", "Status",
            "Actual", "Pred_Adj", "Pred_Adj_Lo", "Pred_Adj_Hi",
            "Actual_Paid", "Pred_Paid", "Pred_Paid_Lo", "Pred_Paid_Hi",
            "Actual_Comp", "Pred_Comp", "Pred_Comp_Lo", "Pred_Comp_Hi",
            "CompRate", "CompRate_Source"]
    headers = cols
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    _style_header(ws, 1, len(headers))

    for i, (_, r) in enumerate(out[cols].iterrows()):
        row = 2 + i
        for c, name in enumerate(cols, start=1):
            v = r[name]
            cell = ws.cell(row=row, column=c, value=(None if pd.isna(v) else v))
            if name == "EventDate":
                cell.number_format = "yyyy-mm-dd"
            elif name == "CompRate":
                cell.number_format = "0.0%"
            elif name in ("Actual", "Pred_Adj", "Pred_Adj_Lo", "Pred_Adj_Hi",
                          "Actual_Paid", "Pred_Paid", "Pred_Paid_Lo", "Pred_Paid_Hi",
                          "Actual_Comp", "Pred_Comp", "Pred_Comp_Lo", "Pred_Comp_Hi",
                          "EventCapacity"):
                cell.number_format = "#,##0"

    last_row = 1 + len(out)
    _stripe(ws, 2, last_row, len(headers))
    ws.freeze_panes = "C2"
    _autowidth(ws, [12, 44, 22, 11, 10, 12, 9, 10, 11, 11, 11, 11, 12, 12, 11, 11, 12, 12, 11, 22])
    ws.sheet_view.showGridLines = False


def main():
    em, merged = load_data()
    prior = sorted([s for s in merged["Season"].dropna().unique()
                    if s < FORECAST_SEASON])
    train = get_training_df(merged, prior)
    (repeat_model, primary_sf, sf_ratio, primary, f1, f2, f3, f3a, f3b, f4, f5
     ) = build_hierarchy_models(train)
    pwyw_lift, _ = build_pwyw_lift(
        merged, prior, repeat_model, primary_sf, sf_ratio,
        primary, f1, f2, f3, f3a, f3b, f4, f5)
    side_lift, side_samples = build_side_lift(
        merged, prior, repeat_model, primary_sf, sf_ratio,
        primary, f1, f2, f3, f3a, f3b, f4, f5)
    print(f"Side-event lift: {side_lift:.2f}x  (n={len(side_samples)} prior-season Side events)")

    # All 25-26 live events from manifest
    em["EventDate"] = pd.to_datetime(em["EventDate"], errors="coerce")
    season = (
        em[(em["Season"] == FORECAST_SEASON)
           & (em["EventType"] == "Live")]
        .drop_duplicates("EventName")
        .copy()
    )
    season["EventCapacity"] = pd.to_numeric(season["EventCapacity"], errors="coerce")

    # Completed actuals — split paid vs comp so the output mirrors Pred_Paid / Pred_Comp
    completed = merged[
        (merged["Season"] == FORECAST_SEASON)
        & (merged["EventType"] == "Live")
        & (merged["EventStatus"] == "Complete")
        & (merged["TicketStatus"] == "Active")
        & (merged["Quantity"] > 0)
    ].copy()
    is_comp = (completed["IsComp"].astype(bool) if "IsComp" in completed.columns
               else completed["TicketTotal"] == 0)
    completed["CompQty"] = np.where(is_comp, completed["Quantity"], 0)
    completed["PaidQty"] = np.where(is_comp, 0, completed["Quantity"])
    actuals = (
        completed.groupby("EventName")
        .agg(Actual=("Quantity", "sum"),
             Actual_Paid=("PaidQty", "sum"),
             Actual_Comp=("CompQty", "sum"))
        .reset_index()
    )
    season = season.merge(actuals, on="EventName", how="left")

    season["Status"] = np.where(season["Actual"].notna(), "Completed", "Upcoming")

    # Predict
    fc = predict_model_a(season, repeat_model, primary_sf, sf_ratio,
                         primary, f1, f2, f3, f3a, f3b, f4, f5,
                         pwyw_lift=pwyw_lift, side_lift=side_lift)
    fc["Pred_A"] = cap_at_capacity(fc["Pred_A"], fc["EventCapacity"])

    # Artist adjustment needs training-season labelled history
    hist_gb = ["EventId", "EventName", "EventClass", "EventVenue",
               "EventGenre", "EventLoB", "EventSubGenre", "EventRepeat"]
    for c in ("SeatFormat", "VenueType"):
        if c in train.columns:
            hist_gb.append(c)
    hist_actuals = (
        train.groupby(hist_gb, group_keys=False, dropna=False)
        .agg(Actual=("Quantity", "sum"))
        .reset_index()
    )
    cap = em.drop_duplicates("EventId")[["EventId", "EventCapacity"]].copy()
    cap["EventCapacity"] = pd.to_numeric(cap["EventCapacity"], errors="coerce")
    hist_actuals = hist_actuals.merge(cap, on="EventId", how="left")
    hist_fc = predict_model_a(hist_actuals, repeat_model, primary_sf, sf_ratio,
                               primary, f1, f2, f3, f3a, f3b, f4, f5)
    fc = apply_artist_adjustment(
        fc,
        merged_history=hist_fc,
        actuals_history=hist_fc["Actual"],
        bucket_preds_history=hist_fc["Pred_A"],
    )

    # Fill CIs for events without artist signal using bucket-level residuals
    fc = apply_bucket_ci(fc)

    # Cap CI bounds at capacity too
    fc["Pred_Adj"]    = cap_at_capacity(fc["Pred_Adj"],    fc["EventCapacity"])
    fc["Pred_Adj_Lo"] = cap_at_capacity(fc["Pred_Adj_Lo"], fc["EventCapacity"])
    fc["Pred_Adj_Hi"] = cap_at_capacity(fc["Pred_Adj_Hi"], fc["EventCapacity"])

    # Paid/comp split — historical comp rate from training seasons, routed
    # by event bucket (EventRepeat → Primary → F1 → F2 → F3 → F4 → overall)
    comp_rates = build_comp_rates(merged, prior)
    fc = apply_comp_split(fc, comp_rates)

    fc["Status"] = np.where(fc["Actual"].notna(), "Completed", "Upcoming")

    out = (
        fc[["EventDate", "EventName", "EventVenue", "EventClass",
            "EventCapacity", "Status",
            "Actual", "Actual_Paid", "Actual_Comp",
            "Pred_Adj", "Pred_Adj_Lo", "Pred_Adj_Hi",
            "Pred_Paid", "Pred_Paid_Lo", "Pred_Paid_Hi",
            "Pred_Comp", "Pred_Comp_Lo", "Pred_Comp_Hi",
            "CompRate", "CompRate_Source"]]
        .sort_values("EventDate")
        .reset_index(drop=True)
    )
    for c in ["Pred_Adj", "Pred_Adj_Lo", "Pred_Adj_Hi",
              "Pred_Paid", "Pred_Paid_Lo", "Pred_Paid_Hi",
              "Pred_Comp", "Pred_Comp_Lo", "Pred_Comp_Hi",
              "Actual", "Actual_Paid", "Actual_Comp", "EventCapacity"]:
        out[c] = out[c].round(0)
    out["CompRate"] = out["CompRate"].round(3)

    # Summary: totals + accuracy for paid+comp, paid only, comp only
    done_df = out[out["Status"] == "Completed"].copy()
    summary_rows = []
    for label, a_col, p_col in [
        ("Total (paid+comp)", "Actual",      "Pred_Adj"),
        ("Paid only",         "Actual_Paid", "Pred_Paid"),
        ("Comp only",         "Actual_Comp", "Pred_Comp"),
    ]:
        sub = done_df[done_df[a_col].notna() & (done_df[a_col] > 0)]
        a, p = sub[a_col].sum(), sub[p_col].sum()
        summary_rows.append({
            "Metric":    label,
            "n":         len(sub),
            "Actual":    int(a),
            "Predicted": int(p),
            "Net Error": int(p - a),
            "% Error":   round((p - a) / a * 100, 2) if a else None,
            "WAPE %":    round((sub[p_col] - sub[a_col]).abs().sum() / a * 100, 1) if a else None,
            "Bias %":    round(((sub[p_col] - sub[a_col]) / sub[a_col]).mean() * 100, 1) if a else None,
        })
    summary = pd.DataFrame(summary_rows)

    completed_only = out[out["Status"] == "Completed"]
    n_total = len(completed_only)
    actual_total = int(completed_only["Actual"].sum())
    pred_total   = int(completed_only["Pred_Adj"].sum())

    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)
    write_summary_sheet(wb.create_sheet("Summary"), summary, n_total, actual_total, pred_total)
    write_business_sheet(wb.create_sheet("Forecast vs Actuals"), out)
    write_detail_sheet(wb.create_sheet("Detail"), out)
    wb.save(OUT_XLSX)

    n_comp = (out["Status"] == "Completed").sum()
    n_upc  = (out["Status"] == "Upcoming").sum()
    done   = out["Actual"].notna()
    upc    = out["Status"] == "Upcoming"
    print(f"✓ {OUT_XLSX}")
    print(f"  Events: {len(out)}  ({n_comp} completed, {n_upc} upcoming)")
    print(f"  Completed to date — total: {int(out.loc[done, 'Actual'].sum()):,}  "
          f"paid: {int(out.loc[done, 'Actual_Paid'].sum()):,}  "
          f"comp: {int(out.loc[done, 'Actual_Comp'].sum()):,}")
    print(f"  Upcoming forecast  — total: {int(out.loc[upc, 'Pred_Adj'].sum()):,}  "
          f"paid: {int(out.loc[upc, 'Pred_Paid'].sum()):,}  "
          f"comp: {int(out.loc[upc, 'Pred_Comp'].sum()):,}")
    print(f"  Full season        — total: {int(out['Pred_Adj'].sum()):,}  "
          f"paid: {int(out['Pred_Paid'].sum()):,}  "
          f"comp: {int(out['Pred_Comp'].sum()):,}")


if __name__ == "__main__":
    main()
