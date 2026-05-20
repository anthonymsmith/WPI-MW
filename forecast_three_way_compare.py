"""
Forecast comparison for the 25-26 season:
 - New hierarchical model (current)
 - The Great Game — pool of seven human guessers (Adrien, Gillian, Heather,
   Kate, Lisa, Rae, Tony Orig)

Outputs a per-event table, a ranked leaderboard, and a summary chart.
"""
import os
import pandas as pd
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

os.chdir("/Users/antho/Documents/WPI-MW")

GG_COLUMNS = ["Adrien", "Gillian", "Heather", "Kate", "Lisa", "Rae", "Tony Orig"]

# Great Game guesses transcribed from "the great game - analysis (1).pdf".
# Event names already mapped to the canonical names used in the forecast file.
GG_DATA = [
    ("Ladysmith Black Mambazo",                                        999,  894,  900,  470,  623,  600,  700),
    ("Emi Ferguson & Ruckus",                                          175,  187,  110,  200,  213,  200,  200),
    ("Hermitage Trio",                                                 231,  526,  257,  250,  241,  389,  300),
    ("TCB: Old Post Road",                                             251,  330,  150,  226,  234,  314,  278),
    ("BACHtoberfest 2025: CONCORA, Baroklyn, Simone Dinnerstein",      601,  941,  800,  325,  316,  680,  336),
    ("BACHtoberfest 2025: Organ",                                      177,  406,  250,  285,  562,  500,  250),
    ("BACHtoberfest 2025: Simone Dinnerstein Recital",                 582,  832,  795,  410,  670,  800,  700),
    ("BACHtoberfest 2025: Choir",                                      303,  361,  400,  250,  433,  400,  336),
    ("Kyung-Wha Chung",                                                729,  636,  942,  600,  814,  800,  920),
    ("Orchestre National de France, Daniil Trifonov",                 1111, 1204, 1075, 1077, 1079, 1100, 1200),
    ("American Patchwork Quartet",                                      99,  216,  239,  300,  222,  330,  336),
    ("Handel Messiah 2025",                                           1205,  848,  841,  825,  910,  759, 1000),
    ("Women's Ensemble Holiday 2025",                                  199,  132,  128,  225,  208,  221,  197),
    ("American Spiritual Ensemble 2026",                               525,  504,  550,  270,  535,  402,  336),
    ("TCB: Christmas Oratorio with Winchendon Players",                608,  420,  600,  491,  270,  345,  449),
    ("TCB: Emmanuel Music: Solo Cantatas",                             179,  461,  270,  255,  194,  360,  214),
    ("Aaron Diehl Trio",                                               273,  214,  290,  280,  281,  380,  421),
    ("Refugee Orchestra Project",                                      489,  702,  891,  550,  590,  780,  575),
    ("Nelson Goerner",                                                 152,  305,  222,  270,  207,  370,  214),
    ("Bach's Birthday Bash 2026: The Sebastians",                      777,  613,  790,  610,  840,  950,  572),
    ("Bach's Birthday Bash 2026: Keyboards Up Close",                   99,  231,  370,  200,  275,  520,  572),
    ("Bach's Birthday Bash 2026: Worcester Chamber Music Society",     333,  510,  441,  340,  349,  461,  400),
    ("Bach's Birthday Bash 2026: Cantatathon",                         492,  329,  400,  350,  435,  500,  432),
    ("Catherine Russell & Sean Mason",                                 222,  317,  250,  220,  284,  270,  200),
    ("Jordi Savall & Hesperion XXI",                                   802,  599,  675,  705,  448,  700,  575),
    ("Alexandre Kantorow: Piano Recital",                              551,  631,  700,  375,  726,  800,  600),
    # Upcoming — no actuals yet, still useful for reporting
    ("Chorus: Frederick Douglass",                                     558,  576,  477,  336,  550,  600,  449),
    ("Women's Ensemble & Cantilena",                                   163,  319,  129,  175,  202,  350,  197),
    ("Dance Theatre of Harlem 2026",                                  1867, 1111, 1100, 2098, 1355, 2104, 1400),
]

NAVY   = "#1A3A5C"
TEAL   = "#2A9EA0"
ORANGE = "#E8922A"
SLATE  = "#8A9BAE"
LGRAY  = "#E8EDF2"
DGRAY  = "#5A6A7A"


def load_all():
    gg = pd.DataFrame(GG_DATA, columns=["EventName"] + GG_COLUMNS)

    comp = pd.read_excel("forecasting/Forecast_2526_FullSeason.xlsx",
                          sheet_name="Detail")
    comp = comp[["EventName", "Actual", "Pred_Adj"]].rename(
        columns={"Pred_Adj": "Model"})

    m = gg.merge(comp, on="EventName", how="left")
    return m


def metrics(df, forecaster):
    sub = df[["Actual", forecaster]].dropna()
    if sub.empty:
        return None
    err = sub[forecaster] - sub["Actual"]
    abs_pct = err.abs() / sub["Actual"] * 100
    signed_pct = err / sub["Actual"] * 100
    mape = abs_pct.mean()
    wape = err.abs().sum() / sub["Actual"].sum() * 100
    bias = signed_pct.mean()
    bias_abs = err.sum() / sub["Actual"].sum() * 100
    mae = err.abs().mean()
    return dict(Forecaster=forecaster, N=len(sub),
                MAPE=mape, WAPE=wape,
                Bias_mean_pct=bias, Bias_total_pct=bias_abs,
                MAE=mae)


def leaderboard(df):
    done = df[df["Actual"].notna()].copy()
    cols = ["Model"] + GG_COLUMNS
    rows = [metrics(done, c) for c in cols]
    lb = pd.DataFrame(rows).sort_values("WAPE").reset_index(drop=True)

    # Wins (closest guess per event), only counting forecasters with a value
    wins = {c: 0 for c in cols}
    for _, r in done.iterrows():
        preds = {c: r[c] for c in cols if pd.notna(r[c])}
        if not preds:
            continue
        winner = min(preds, key=lambda c: abs(preds[c] - r["Actual"]))
        wins[winner] += 1
    lb["Wins"] = lb["Forecaster"].map(wins)
    return lb, done


def plot_leaderboard(lb, out="forecasting/forecast_three_way_leaderboard.png"):
    fig, ax = plt.subplots(figsize=(10.5, 5.6))
    order = lb.sort_values("WAPE", ascending=False)

    def color(name):
        if name == "Model": return NAVY
        return TEAL

    colors = [color(n) for n in order["Forecaster"]]
    y = np.arange(len(order))
    ax.barh(y, order["WAPE"], color=colors, edgecolor="white", linewidth=1.0, zorder=3)

    for yi, name, wape, bias_tot in zip(
            y, order["Forecaster"], order["WAPE"], order["Bias_total_pct"]):
        ax.text(wape + 0.8, yi,
                f"{wape:.0f}%  ·  bias {bias_tot:+.0f}%",
                va="center", fontsize=10, color=DGRAY)

    labels = ["New model" if n == "Model" else n
              for n in order["Forecaster"]]
    ax.set_yticks(y); ax.set_yticklabels(labels, fontsize=11)
    ax.set_xlim(0, max(order["WAPE"]) * 1.35)
    ax.set_xlabel("Season-total error  (WAPE %)", color=DGRAY)
    ax.grid(axis="x", color=LGRAY, linewidth=0.7, zorder=0)
    ax.set_axisbelow(True)
    for spine in ("top", "right"):
        ax.spines[spine].set_visible(False)
    ax.spines["left"].set_color(LGRAY); ax.spines["bottom"].set_color(LGRAY)

    fig.suptitle("25–26 Season — Budget-Relevant Forecast Accuracy",
                 fontsize=14, fontweight="bold", color=NAVY,
                 x=0.02, ha="left", y=0.98)
    fig.text(0.02, 0.93,
             "WAPE = season-total error (lower = better)  ·  "
             "Bias = season-total direction (closer to 0 = better)",
             ha="left", va="top", fontsize=10, color=DGRAY)
    fig.tight_layout(rect=[0, 0, 1, 0.92])
    fig.savefig(out, dpi=180, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    print(f"  ✓ {out}")


def main():
    m = load_all()
    lb, done = leaderboard(m)

    print("\n=== Leaderboard (completed events only) ===")
    print(lb.to_string(index=False, float_format=lambda v: f"{v:.1f}"))

    print(f"\nEvents in comparison: {len(m)}  "
          f"({len(done)} completed, {len(m) - len(done)} upcoming)")
    print("\n=== Per-event detail ===")
    show = m[["EventName", "Model"] + GG_COLUMNS + ["Actual"]].copy()
    for c in ["Model"] + GG_COLUMNS:
        show[c] = show[c].apply(lambda v: f"{v:.0f}" if pd.notna(v) else "")
    show["Actual"] = show["Actual"].apply(
        lambda v: f"{int(v)}" if pd.notna(v) else "—")
    print(show.to_string(index=False))

    plot_leaderboard(lb)

    write_formatted_xlsx(m, lb)


def write_formatted_xlsx(m, lb, out_xlsx="Forecast 3-way formatted.xlsx"):
    forecaster_cols = GG_COLUMNS + ["Model"]

    # AllEvents: EventName, 7 humans, Model, Actual
    col_order = ["EventName"] + GG_COLUMNS + ["Model", "Actual"]
    ae = m[col_order].copy()

    # Leaderboard: trimmed/renamed, sorted by Wins desc then WAPE asc
    lb_out = lb.rename(columns={
        "WAPE":           "WAPE (Total Abs % Error)",
        "Bias_total_pct": "% Bias",
        "MAE":            "Mean Absolute Error",
    })[["Forecaster", "WAPE (Total Abs % Error)", "% Bias",
        "Mean Absolute Error", "Wins"]]
    lb_out = lb_out.sort_values(
        ["Wins", "WAPE (Total Abs % Error)"], ascending=[False, True]
    ).reset_index(drop=True)

    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as w:
        lb_out.to_excel(w, sheet_name="Leaderboard", index=False)
        ae.to_excel(w, sheet_name="AllEvents", index=False)

    # --- formatting pass ---
    wb = load_workbook(out_xlsx)
    bold = Font(bold=True)
    red  = Font(color="FFFF0000")

    # AllEvents
    ws = wb["AllEvents"]
    for c in range(1, ws.max_column + 1):
        ws.cell(1, c).font = bold
    ws.column_dimensions["A"].width = 45
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    for col_name in headers[1:]:
        letter = get_column_letter(headers.index(col_name) + 1)
        ws.column_dimensions[letter].width = 8.8 if col_name == "Actual" else 13

    actual_c = headers.index("Actual") + 1
    for r in range(2, ws.max_row + 1):
        actual = ws.cell(r, actual_c).value
        if actual is None:
            continue
        preds = {col: ws.cell(r, headers.index(col) + 1).value
                 for col in forecaster_cols
                 if ws.cell(r, headers.index(col) + 1).value is not None}
        if not preds:
            continue
        best_err = min(abs(v - actual) for v in preds.values())
        for col, v in preds.items():
            if abs(v - actual) == best_err:
                ws.cell(r, headers.index(col) + 1).font = red

    # Leaderboard
    ws = wb["Leaderboard"]
    for c in range(1, ws.max_column + 1):
        ws.cell(1, c).font = bold
    widths = {"A": 12, "B": 24, "C": 10, "D": 22, "E": 7}
    for letter, w in widths.items():
        ws.column_dimensions[letter].width = w
    for r in range(2, ws.max_row + 1):
        for c in (2, 3, 4):
            ws.cell(r, c).number_format = "0.0"

    wb.save(out_xlsx)
    print(f"  ✓ {out_xlsx}")


if __name__ == "__main__":
    main()
